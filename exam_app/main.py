"""
Main application module for ExamScheduler.

Contains FastAPI endpoints, Google authentication, file processing,
and routing for dashboard and user interface views.
"""
import os
import io
import time
import asyncio
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any

from dotenv import load_dotenv
load_dotenv()

import jwt
import pandas as pd
import aiohttp
import requests as http_requests
from openpyxl import Workbook
from fpdf import FPDF

from fastapi import (
    FastAPI, Request, Depends, HTTPException,
    UploadFile, File, Form, Query, Body
)
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import (
    HTMLResponse, RedirectResponse, StreamingResponse, 
    FileResponse, JSONResponse
)
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware

from google.oauth2 import id_token
from google.auth.transport.requests import Request as GoogleRequest

from sqlalchemy import select, text, or_, and_, func, case, String
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.postgresql import insert as pg_insert

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Local module imports
from models import (
    Faculty,
    Professor,
    Secretariat,
    Admin,
    GroupLeader,
    Subject,
    Subgroup,
    ExamProposal,
    Room,
    ExamLimits
)
from database import SessionLocal, engine, Base
from schemas import ProposalCreate

# Register custom fonts for PDF generation
pdfmetrics.registerFont(TTFont('DejaVuSans', 'static/fonts/DejaVuSans.ttf'))

app = FastAPI(title="ExamScheduler System")

templates = Jinja2Templates(directory="templates")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# JWT & Google Security Configurations
SECRET_KEY = os.getenv("SECRET_KEY", "default-strong-secret-key")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "default_google_client_id")

# Create database tables if they do not exist
Base.metadata.create_all(bind=engine)

security = HTTPBearer()


# ==========================================
# HELPER FUNCTIONS & DEPENDENCIES
# ==========================================

async def fetch_json_with_timeout(session: aiohttp.ClientSession, url: str, timeout_sec: int = 7) -> Dict[str, Any]:
    """Fetches JSON payload asynchronously with a strict timeout."""
    try:
        async with session.get(url, timeout=timeout_sec) as response:
            response.raise_for_status()
            return await response.json()
    except Exception as err:
        raise err


def get_db():
    """Provides a transactional database session scope per request."""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    """Generates a signed JWT access token."""
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


async def get_current_user(
    request: Request,
    db: Session = Depends(get_db),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False))
) -> Dict[str, Any]:
    """Validates the JWT token from authorization headers or cookies and yields the active user."""
    token = credentials.credentials if credentials else request.cookies.get("access_token")

    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        role: str = payload.get("role")
        if not email or not role:
            raise HTTPException(status_code=403, detail="Invalid authorization token")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=403, detail="Authorization token expired")
    except jwt.PyJWTError:
        raise HTTPException(status_code=403, detail="Invalid or corrupt token")

    # Map application roles to database models
    role_model_mapping = {
        "admin": Admin,
        "secretariat": Secretariat,
        "professor": Professor,
        "group_leader": GroupLeader,
    }

    model = role_model_mapping.get(role)
    if not model:
        raise HTTPException(status_code=403, detail="Unknown system role")

    user = db.query(model).filter_by(emailAddress=email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User account not found")

    return {
        "id": user.id,
        "email": email,
        "role": role
    }


# ==========================================
# AUTHENTICATION & CORE ENDPOINTS
# ==========================================

@app.post("/token")
async def handle_google_token(request: Request, db: Session = Depends(get_db)):
    """Verifies Google OAuth2 ID token and issues system JWT cookie."""
    data = await request.json()
    google_token = data.get("token")

    if not google_token:
        raise HTTPException(status_code=400, detail="Missing Google Token")

    try:
        id_info = id_token.verify_oauth2_token(
            google_token,
            GoogleRequest(),
            GOOGLE_CLIENT_ID
        )
        email = id_info.get("email")

        # Resolve user role across database entities
        professor = db.query(Professor).filter(Professor.emailAddress == email).first()
        secretariat_member = db.query(Secretariat).filter(Secretariat.emailAddress == email).first()
        admin_member = db.query(Admin).filter(Admin.emailAddress == email).first()
        leader_member = db.query(GroupLeader).filter(GroupLeader.emailAddress == email).first()

        if professor:
            role = "professor"
        elif secretariat_member:
            role = "secretariat"
        elif admin_member:
            role = "admin"
        elif leader_member:
            role = "group_leader"
        else:
            raise HTTPException(status_code=403, detail="Email is not registered in the system")

        token_payload = {
            "sub": email,
            "role": role
        }
        access_token = create_access_token(data=token_payload)
        response = JSONResponse(content={"role": role})
        response.set_cookie(key="access_token", value=access_token, httponly=True)
        return response

    except ValueError as val_err:
        raise HTTPException(status_code=401, detail=f"Invalid Google Token: {str(val_err)}")
    except Exception as sys_err:
        raise HTTPException(status_code=500, detail=f"Internal Server Error: {str(sys_err)}")


@app.get("/dashboard")
async def dashboard(current_user: dict = Depends(get_current_user)):
    """Protected dashboard root route."""
    return {"message": f"Welcome back, {current_user['email']}!"}


@app.get("/users/me")
async def read_users_me(current_user: dict = Depends(get_current_user)):
    """Returns details about the currently authenticated user."""
    return current_user


@app.get("/server-time")
async def get_server_time():
    """Returns current system UTC and local server time."""
    return {
        "timestamp": int(time.time()),
        "datetime": str(datetime.utcnow()),
        "timezone": str(datetime.now().astimezone().tzinfo)
    }


@app.get("/faculties/")
def list_faculties(db: Session = Depends(get_db)):
    """Lists all faculties available in the database."""
    return db.query(Faculty).all()


@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Session = Depends(get_db)):
    """Renders the primary application homepage upon session validation."""
    token = request.cookies.get("access_token")
    if not token:
        return RedirectResponse(url="/login")

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        role = payload.get("role")
    except Exception:
        return RedirectResponse(url="/login")

    user_data = None
    for model in [Professor, Secretariat, Admin, GroupLeader]:
        person = db.query(model).filter(model.emailAddress == email).first()
        if person:
            user_data = {
                "full_name": f"{person.firstName} {person.lastName}",
                "email": email,
                "role": role
            }
            break

    if not user_data:
        return RedirectResponse(url="/login")

    return templates.TemplateResponse("home.html", {
        "request": request,
        "user": user_data,
        "full_name": user_data["full_name"]
    })


@app.get("/logout")
async def logout():
    """Clears authentication token cookies and redirects to login."""
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("access_token")
    return response


@app.post("/logout")
async def logout():
    response = JSONResponse(
        content={"message": "Logged out successfully"}
    )

    response.delete_cookie(
        key="access_token"
    )

    return response

# ==========================================
# SECRETARIAT & GROUP LEADER WORKFLOWS
# ==========================================

@app.get("/upload-Subject", response_class=HTMLResponse)
async def upload_page(
    request: Request, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Renders file upload portal for Subjectingestion. Secretariat access only."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    return templates.TemplateResponse("upload_Subject.html", {
        "request": request
    })


@app.get("/proposals")
async def view_group_leader_subjects(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user["role"] != "group_leader":
        raise HTTPException(
            status_code=403,
            detail="Access restricted to Group Leaders only"
        )

    leader = (
        db.query(GroupLeader)
        .filter(GroupLeader.emailAddress == current_user["email"])
        .first()
    )

    if not leader:
        raise HTTPException(
            status_code=404,
            detail="Group Leader not found"
        )

    subjects = (
        db.query(Subject)
        .filter(Subject.subgroup_id == leader.subgroup_id)
        .all()
    )

    subgroup = leader.subgroup

    return {
        "group": (
            f"{subgroup.groupName}{subgroup.subgroupIndex}"
            if subgroup else None
        ),
        "Subjects": [
            {
                "id": subject.id,
                "name": subject.topic,
                "tenured_professor": (
                    f"{subject.professor.firstName} "
                    f"{subject.professor.lastName}"
                    if subject.professor else None
                ),
                "email": (
                    subject.professor.emailAddress
                    if subject.professor else None
                )
            }
            for subject in subjects
        ]
    }

@app.get("/group-leader/Subjects")
async def list_available_Subjects(
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Lists group Subjects that have not yet been submitted or accepted."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter(GroupLeader.emailAddress == current_user["email"]).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group Leader entity not found")

    subquery = db.query(ExamProposal.subject_id).filter(
        ExamProposal.status.in_(["submitted", "accepted"])
    ).subquery()

    subjects = db.query(Subject).filter(
        Subject.subgroup_id == leader.subgroup_id,
        ~Subject.id.in_(select(subquery))
    ).all()

    return [
        {
            "id": d.id, 
            "topic": d.topic, 
            "faculty_member": ( f"{d.professor.firstName} {d.professor.lastName}" if d.professor else "" )
        } for d in subjects
    ]


@app.get("/group-leader/proposal", response_class=HTMLResponse)
async def display_proposal_form(
    request: Request, 
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Displays exam proposal submission view for Group Leaders."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter(GroupLeader.emailAddress == current_user["email"]).first()

    subquery = db.query(ExamProposal.subject_id).filter(
        ExamProposal.status.in_(["submitted", "accepted"])
    ).subquery()

    Subjects = db.query(Subject).filter(
        Subject.subgroup_id == leader.subgroup_id,
        ~Subject.id.in_(subquery)
    ).all()

    return templates.TemplateResponse("group_leader_proposal.html", {
        "request": request,
        "Subjects": Subjects
    })


@app.get("/group-leader/proposal-form", response_class=HTMLResponse)
async def proposal_selection_form(
    request: Request, 
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Renders alternative Subjectselection view."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter(GroupLeader.emailAddress == current_user["email"]).first()

    subquery = db.query(ExamProposal.subject_id).filter(
        ExamProposal.status.in_(["submitted", "accepted"])
    ).subquery()

    Subjects = db.query(Subject).filter(
        Subject.subgroup_id == leader.subgroup_id,
        ~Subject.id.in_(subquery)
    ).all()

    return templates.TemplateResponse("select_Subject.html", {
        "request": request,
        "Subjects": Subjects
    })


@app.post("/group-leader/proposal")
async def submit_proposal(
    proposal: ProposalCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Submits or resubmits an exam date proposal."""

    if current_user["role"] != "group_leader":
        raise HTTPException(
            status_code=403,
            detail="Access restricted to group leaders"
        )

    leader = db.query(GroupLeader).filter(
        GroupLeader.emailAddress == current_user["email"]
    ).first()

    if not leader:
        raise HTTPException(
            status_code=404,
            detail="Group leader not found"
        )

    subject = db.query(Subject).filter(
        Subject.id == proposal.subject_id,
        Subject.subgroup_id == leader.subgroup_id
    ).first()

    if not subject:
        raise HTTPException(
            status_code=403,
            detail="Subject does not belong to this group"
        )

    start_time = proposal.date.replace(tzinfo=None)
    end_time = start_time + timedelta(hours=proposal.duration)

    existing_proposals = (
        db.query(ExamProposal)
        .join(Subject)
        .filter(
            Subject.subgroup_id == leader.subgroup_id,
            ExamProposal.status.in_(["submitted", "accepted"])
        )
        .all()
    )

    for existing in existing_proposals:
        existing_start = existing.date.replace(tzinfo=None)
        existing_end = existing_start + timedelta(
            hours=existing.duration
        )

        if (
            start_time < existing_end
            and end_time > existing_start
        ):
            raise HTTPException(
                status_code=409,
                detail=(
                    "Proposed time slot overlaps with "
                    "another exam for this group."
                )
            )

    rejected_proposal = (
        db.query(ExamProposal)
        .filter(
            ExamProposal.subject_id == proposal.subject_id,
            ExamProposal.group_leader_id == leader.id,
            ExamProposal.status == "rejected"
        )
        .order_by(ExamProposal.id.desc())
        .first()
    )

    if rejected_proposal:
        rejected_proposal.date = proposal.date
        rejected_proposal.duration = proposal.duration
        rejected_proposal.status = "submitted"
        rejected_proposal.rejection_reason = None
        rejected_proposal.room_id = None
        rejected_proposal.assistant_id = None
    else:
        new_proposal = ExamProposal(
            subject_id=proposal.subject_id,
            group_leader_id=leader.id,
            date=proposal.date,
            duration=proposal.duration,
            status="submitted"
        )

        db.add(new_proposal)

    db.commit()

    return {
        "message": "Proposal submitted successfully"
    }


@app.get("/group-leader/calendar-occupancy")
async def calendar_occupancy(
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Retrieves schedule occupancy data for group calendar rendering."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter(GroupLeader.emailAddress == current_user["email"]).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group Leader entity not found")

    proposals = db.query(ExamProposal).join(Subject).filter(
        Subject.subgroup_id == leader.subgroup_id
    ).all()

    return [
        {
            "date": p.date.isoformat(),
            "duration": p.duration,
            "status": p.status,
            "rejection_reason": p.rejection_reason,
            "subject": p.subject.topic
        }
        for p in proposals
    ]


@app.get("/group-leader/calendar", response_class=HTMLResponse)
async def display_calendar(
    request: Request,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Renders calendar user interface populated with session date bounds."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    limits = db.query(ExamLimits).first()

    return templates.TemplateResponse("calendar.html", {
        "request": request,
        "start_limit": limits.start_date.isoformat() if limits else "",
        "end_limit": limits.end_date.isoformat() if limits else ""
    })


# ==========================================
# FACULTY MEMBER WORKFLOWS
# ==========================================

@app.get("/professor/proposals", response_class=HTMLResponse)
async def display_professor_proposals(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Displays incoming proposals for review by the logged-in faculty member."""
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Access restricted to Faculty Members only")

    professor = db.query(Professor).filter(Professor.emailAddress == current_user["email"]).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    proposals = db.query(ExamProposal).join(Subject).filter(
        Subject.professor_id == professor.id,
        ExamProposal.status == "submitted"
    ).all()

    all_rooms= db.query(Room).all()
    department_colleagues = db.query(Professor).filter(
        Professor.departmentName == professor.departmentName,
        Professor.id != professor.id
    ).all()

    available_rooms= {}
    available_assistants = {}

    for prop in proposals:
        start_time = prop.date.replace(tzinfo=None)
        end_time = start_time + timedelta(hours=prop.duration)

        occupied_exams = db.query(ExamProposal).filter(
            ExamProposal.room_id.isnot(None),
            ExamProposal.status == "accepted"
        ).all()

        occupied_room_ids = set()
        for occupied in occupied_exams:
            p_start = occupied.date.replace(tzinfo=None)
            p_end = p_start + timedelta(hours=occupied.duration)
            if start_time < p_end and end_time > p_start:
                occupied_room_ids.add(occupied.room_id)

        available_rooms[prop.id] = [room for room in all_rooms if room.id not in occupied_room_ids]

        eligible_assistants = []
        for colleague in department_colleagues:
            assistant_exams = db.query(ExamProposal).filter(
                ExamProposal.assistant_id == colleague.id,
                ExamProposal.status == "accepted"
            ).all()

            has_conflict = False
            for assistant_exam in assistant_exams:
                ep_start = assistant_exam.date.replace(tzinfo=None)
                ep_end = ep_start + timedelta(hours=assistant_exam.duration)
                if start_time < ep_end and end_time > ep_start:
                    has_conflict = True
                    break

            if not has_conflict:
                eligible_assistants.append(colleague)

        available_assistants[prop.id] = eligible_assistants

    return templates.TemplateResponse("professor_proposals.html", {
        "request": request,
        "proposals": proposals,
        "professor": professor,
        "available_rooms": available_rooms,
        "available_assistants": available_assistants
    })


@app.post("/professor/proposals/accept")
async def accept_proposal(
    proposal_id: int = Form(...),
    room_id: int = Form(...),
    assistant_id: int = Form(...),
    duration: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Accepts an exam proposal and assigns the specified room and teaching assistant."""
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Access restricted to Faculty Members only")

    professor = db.query(Professor).filter(Professor.emailAddress == current_user["email"]).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    proposal = db.query(ExamProposal).filter(ExamProposal.id == proposal_id).first()
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal does not exist")

    subject= db.query(Subject).filter(Subject.id == proposal.subject_id).first()
    if not subject or subject.professor_id != professor.id:
        raise HTTPException(status_code=403, detail="You are not authorized for this Subject")

    proposal.status = "accepted"
    proposal.room_id = room_id
    proposal.assistant_id = assistant_id
    proposal.duration = duration

    db.commit()

    return {
        "message": "Proposal accepted successfully"
    }


@app.post("/professor/proposals/reject")
async def reject_proposal(
    proposal_id: int = Form(...),
    reason: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Rejects an exam proposal with a provided explanatory reason."""

    if current_user["role"] != "professor":
        raise HTTPException(
            status_code=403,
            detail="Access restricted to professors only"
        )

    professor = db.query(Professor).filter(
        Professor.emailAddress == current_user["email"]
    ).first()

    if not professor:
        raise HTTPException(
            status_code=404,
            detail="Professor not found"
        )

    proposal = db.query(ExamProposal).filter(
        ExamProposal.id == proposal_id
    ).first()

    if not proposal:
        raise HTTPException(
            status_code=404,
            detail="Proposal not found"
        )

    subject = db.query(Subject).filter(
        Subject.id == proposal.subject_id
    ).first()

    if not subject or subject.professor_id != professor.id:
        raise HTTPException(
            status_code=403,
            detail="Not authorized to reject this proposal"
        )

    proposal.status = "rejected"
    proposal.rejection_reason = reason

    db.commit()

    return {
        "message": "Proposal rejected successfully"
    }


# ==========================================
# FACULTY & GROUP LEADER EXTRA ACTIONS
# ==========================================

@app.post("/professor/proposal/{id}/reject")
async def reject_exam_proposal(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Rejects a specific exam proposal by ID."""
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Only faculty members can reject proposals")

    faculty = db.query(Professor).filter(Professor.emailAddress == current_user["email"]).first()
    proposal = db.query(ExamProposal).filter_by(id=id).first()

    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal does not exist")

    subject= db.query(Subject).filter_by(id=proposal.subject_id).first()
    if not subject or subject.professor_id != faculty.id:
        raise HTTPException(status_code=403, detail="You are not the tenured professor for this Subject")

    proposal.status = "rejected"
    db.commit()

    return RedirectResponse(url="/professor/proposals", status_code=303)


@app.get("/professor/accepted-exams", response_class=HTMLResponse)
async def list_accepted_faculty_exams(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Lists all accepted exams where the user is the main course professor."""
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Access restricted to Faculty Members only")

    professor = db.query(Professor).filter(Professor.emailAddress == current_user["email"]).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    proposals = db.query(ExamProposal).join(Subject).filter(
        Subject.professor_id == professor.id,
        ExamProposal.status == "accepted"
    ).all()

    return templates.TemplateResponse("professor_accepted_exams.html", {
        "request": request,
        "professor": professor,
        "exams": proposals
    })


@app.get("/group-leader/Subject-status", response_class=HTMLResponse)
async def group_leader_Subject_status(
    request: Request,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Overview of Subjectstatus categorizing submitted, accepted, rejected, and unsubmitted exams."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter(GroupLeader.emailAddress == current_user["email"]).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group Leader entity not found")

    Subjects = db.query(Subject).filter(Subject.subgroup_id == leader.subgroup_id).all()
    proposals = db.query(ExamProposal).filter(ExamProposal.group_leader_id == leader.id).all()

    proposals_map = {p.subject_id: p for p in proposals}

    submitted, accepted, rejected, pending = [], [], [], []

    for d in Subjects:
        prop = proposals_map.get(d.id)
        subgroup = d.subgroup

        info = {
            "Subject": d.topic,
            "year": subgroup.studyYear,
            "group": f"{subgroup.groupName}{subgroup.subgroupIndex}",
        }

        if prop:
            info.update({
                "date": prop.date,
                "duration": prop.duration,
                "status": prop.status,
                "reason": prop.rejection_reason
            })
            if prop.status == "submitted":
                submitted.append(info)
            elif prop.status == "accepted":
                accepted.append(info)
            elif prop.status == "rejected":
                rejected.append(info)
        else:
            pending.append(info)

    return templates.TemplateResponse("group_leader_Subject_status.html", {
        "request": request,
        "submitted": submitted,
        "accepted": accepted,
        "rejected": rejected,
        "pending": pending,
        "leader": leader
    })


@app.get("/group-leader/export-excel")
async def export_group_leader_excel(
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Exports status summary of all group Subjectexams to an Excel spreadsheet."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter_by(emailAddress=current_user["email"]).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group Leader entity not found")

    Subjects = db.query(Subject).filter_by(subgroup_id=leader.subgroup_id).all()

    proposals_map = {
        p.subject_id: p
        for p in db.query(ExamProposal).join(Subject).filter(Subject.subgroup_id == leader.subgroup_id).all()
    }

    data = []
    for d in Subjects:
        subgroup = d.subgroup
        group_str = f"{subgroup.studyYear} {subgroup.groupName}{subgroup.subgroupIndex}"
        p = proposals_map.get(d.id)

        data.append({
            "Subject": d.topic,
            "Status": p.status if p else "not_submitted",
            "Date": p.date.strftime("%Y-%m-%d %H:%M") if p else "",
            "Duration (h)": p.duration if p else "",
            "Group": group_str,
            "Rejection Reason": p.rejection_reason if p and p.rejection_reason else "",
            "Room": p.room.name if p and p.room else "",
            "Assistant": f"{p.assistant.firstName} {p.assistant.lastName}" if p and p.assistant else ""
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Exams")

    output.seek(0)
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=group_leader_exams.xlsx"}
    )


@app.get("/professor/assistant-exams", response_class=HTMLResponse)
async def list_assistant_exams(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Displays exams where the faculty member is assigned as a Teaching Assistant."""
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Access restricted to Faculty Members only")

    professor = db.query(Professor).filter_by(emailAddress=current_user["email"]).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    proposals = db.query(ExamProposal).join(Subject).join(Subgroup).filter(
        ExamProposal.status == "accepted",
        ExamProposal.assistant_id == professor.id
    ).all()

    return templates.TemplateResponse("professor_assistant_exams.html", {
        "request": request,
        "professor": professor,
        "exams": proposals
    })


@app.get("/professor/export-excel")
async def export_professor_exams_excel(
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Exports all primary and teaching assistant exam schedules for a faculty member."""
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Access restricted to Faculty Members only")

    professor = db.query(Professor).filter_by(emailAddress=current_user["email"]).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    wb = Workbook()
    primary_ws = wb.active
    primary_ws.title = "Primary Professor Exams"
    primary_ws.append(["Subject", "Year", "Group", "Date & Time", "Duration", "Room", "Assistant"])

    primary_exams = db.query(ExamProposal).join(Subject).filter(
        Subject.professor_id == professor.id,
        ExamProposal.status == "accepted"
    ).all()

    for p in primary_exams:
        sub = p.subject.subgroup
        group = f"{sub.studyYear} {sub.groupName}{sub.subgroupIndex}"
        room = p.room.name if p.room else "-"
        assistant = f"{p.assistant.firstName} {p.assistant.lastName}" if p.assistant else "-"
        primary_ws.append([
            p.subject.topic, sub.studyYear, group, 
            p.date.strftime("%Y-%m-%d %H:%M"), p.duration, room, assistant
        ])

    assistant_ws = wb.create_sheet("Assistant Exams")
    assistant_ws.append(["Subject", "Year", "Group", "Date & Time", "Duration", "Room", "Primary Professor"])

    assistant_exams = db.query(ExamProposal).join(Subject).filter(
        ExamProposal.assistant_id == professor.id,
        ExamProposal.status == "accepted"
    ).all()

    for p in assistant_exams:
        sub = p.subject.subgroup
        group = f"{sub.studyYear} {sub.groupName}{sub.subgroupIndex}"
        room = p.room.name if p.room else "-"
        professor = f"{p.subject.professor.firstName} {p.subject.professor.lastName}"
        assistant_ws.append([
            p.subject.topic, sub.studyYear, group, 
            p.date.strftime("%Y-%m-%d %H:%M"), p.duration, room, professor
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=professor_exams.xlsx"}
    )


# ==========================================
# ADMINISTRATIVE MANAGEMENT WORKFLOWS
# ==========================================

@app.get("/admin/Faculty", response_class=HTMLResponse)
async def list_admin_Faculty(
    request: Request, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Renders faculty administration interface."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    faculties = db.query(Faculty).all()
    return templates.TemplateResponse("admin_Faculty.html", {
        "request": request,
        "Faculty": faculties
    })


@app.post("/admin/Faculty/update")
async def update_Faculty(
    request: Request, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Batch updates existing faculty metadata."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    form = await request.form()
    faculties = db.query(Faculty).all()

    for f in faculties:
        new_long_name = form.get(f"name_{f.id}")
        new_short_name = form.get(f"short_{f.id}")
        if new_long_name and new_short_name:
            f.longName = new_long_name.strip()
            f.shortName = new_short_name.strip()

    db.commit()
    return RedirectResponse("/admin/Faculty", status_code=303)


@app.post("/admin/Faculty/add")
async def add_faculty(
    request: Request, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Creates a new faculty entry."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    form = await request.form()
    name = form.get("longName", "").strip()
    short = form.get("shortName", "").strip()

    if not name or not short:
        raise HTTPException(status_code=400, detail="Missing required field data")

    db.execute(text("SELECT setval('faculties_id_seq', (SELECT MAX(id) FROM faculties))"))
    new_faculty = Faculty(longName=name, shortName=short)
    db.add(new_faculty)
    db.commit()

    return RedirectResponse("/admin/Faculty", status_code=303)


@app.post("/admin/Faculty/delete")
async def delete_faculty(
    request: Request, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Deletes an unreferenced faculty entry."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    form = await request.form()
    faculty_id = form.get("id")

    faculty = db.query(Faculty).filter(Faculty.id == faculty_id).first()
    if not faculty:
        raise HTTPException(status_code=404, detail="Faculty record not found")

    subgroup_count = db.query(Subgroup).filter(Subgroup.facultyId == faculty_id).count()
    if subgroup_count > 0:
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete faculty because related Subgroup still exist."
        )

    db.delete(faculty)
    db.commit()
    return RedirectResponse("/admin/Faculty", status_code=303)


@app.post("/admin/sync-Faculty-sequence")
async def sync_Faculty_sequence(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Resynchronizes the PostgreSQL primary key sequence for the Faculty table."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access restricted to administrators")

    try:
        db.execute(text("SELECT setval('faculties_id_seq', (SELECT COALESCE(MAX(id), 1) FROM faculties))"))
        db.commit()
        return {"message": "Faculty ID sequence successfully resynchronized!"}
    except Exception as err:
        raise HTTPException(status_code=500, detail=f"Sequence sync error: {str(err)}")


@app.get("/admin/faculty-members", response_class=HTMLResponse)
async def admin_faculty_members_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Renders faculty member administration directory."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    faculty_members = db.query(Professor).filter(Professor.emailAddress.ilike('%@usm.ro')).all()

    return templates.TemplateResponse("admin_faculty_members.html", {
        "request": request,
        "faculty_members": faculty_members
    })


@app.post("/admin/faculty-members/update")
async def update_faculty_member(
    id: int = Form(...),
    firstName: str = Form(...),
    lastName: str = Form(...),
    emailAddress: str = Form(...),
    phoneNumber: Optional[str] = Form(None),
    facultyName: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Updates profile attributes for a specific faculty member."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    member = db.query(Professor).filter(Professor.id == id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Professor not found")

    member.firstName = firstName.strip()
    member.lastName = lastName.strip()
    member.emailAddress = emailAddress.strip()
    member.phoneNumber = phoneNumber.strip() if phoneNumber else None
    member.facultyName = facultyName.strip()
    member.departmentName = departmentName.strip()
    db.commit()

    return RedirectResponse("/admin/faculty-members", status_code=303)


@app.post("/admin/faculty-members/delete/{id}")
async def delete_faculty_member(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Removes a faculty member entry from the system."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    member = db.query(Professor).filter(Professor.id == id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Professor not found")

    db.delete(member)
    db.commit()

    return RedirectResponse("/admin/faculty-members", status_code=303)


@app.get("/admin/add-faculty-member", response_class=HTMLResponse)
async def add_faculty_member_form(
    request: Request, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Renders form to create a new faculty member."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    faculties = db.query(Faculty).all()
    departments = db.query(Professor.departmentName).distinct().all()
    department_list = [d[0] for d in departments if d[0]]

    return templates.TemplateResponse("add_faculty_member.html", {
        "request": request,
        "Faculty": faculties,
        "departments": department_list
    })


@app.post("/admin/add-faculty-member")
async def add_faculty_member(
    request: Request,
    lastName: str = Form(...),
    firstName: str = Form(...),
    emailAddress: str = Form(...),
    phoneNumber: Optional[str] = Form(None),
    facultyName: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Persists a new faculty member to the database."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    if not emailAddress.endswith("@usm.ro"):
        raise HTTPException(status_code=400, detail="Email domain must match @usm.ro")

    if db.query(Professor).filter_by(emailAddress=emailAddress).first():
        raise HTTPException(status_code=400, detail="Email address is already registered")

    db.execute(text("SELECT setval('professors_id_seq', (SELECT COALESCE(MAX(id), 1) FROM professors))"))

    new_member = Professor(
        lastName=lastName.strip(),
        firstName=firstName.strip(),
        emailAddress=emailAddress.strip().lower(),
        phoneNumber=phoneNumber.strip() if phoneNumber else None,
        facultyName=facultyName.strip(),
        departmentName=departmentName.strip()
    )

    db.add(new_member)
    db.commit()
    return RedirectResponse("/admin/faculty-members", status_code=303)


@app.get("/api/departments")
async def get_departments_by_faculty(faculty: str, db: Session = Depends(get_db)):
    """API endpoint returning distinct department names for a specified faculty."""
    departments = db.query(Professor.departmentName).filter(
        Professor.facultyName == faculty
    ).distinct().all()
    return [d[0] for d in departments if d[0]]


@app.get("/admin/secretariat", response_class=HTMLResponse)
async def admin_secretariat_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Renders secretariat administration directory."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    secretaries = db.query(Secretariat).order_by(Secretariat.lastName, Secretariat.firstName).all()

    return templates.TemplateResponse("admin_secretariat.html", {
        "request": request,
        "secretaries": secretaries
    })


@app.post("/admin/secretariat/add")
async def add_secretariat_member(
    firstName: str = Form(...),
    lastName: str = Form(...),
    emailAddress: str = Form(...),
    facultyName: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Adds a new secretariat user account."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    if db.query(Secretariat).filter_by(emailAddress=emailAddress).first():
        raise HTTPException(status_code=400, detail="Email address is already registered")

    new_secretary = Secretariat(
        firstName=firstName.strip(),
        lastName=lastName.strip(),
        emailAddress=emailAddress.strip().lower(),
        facultyName=facultyName.strip(),
        departmentName=departmentName.strip()
    )

    db.add(new_secretary)
    db.commit()
    return RedirectResponse("/admin/secretariat", status_code=303)


@app.post("/admin/secretariat/update")
async def update_secretariat_member(
    id: int = Form(...),
    firstName: str = Form(...),
    lastName: str = Form(...),
    emailAddress: str = Form(...),
    facultyName: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Updates attributes of an existing secretariat user account."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    secretary = db.query(Secretariat).filter_by(id=id).first()
    if not secretary:
        raise HTTPException(status_code=404, detail="Secretariat record does not exist")

    secretary.firstName = firstName.strip()
    secretary.lastName = lastName.strip()
    secretary.emailAddress = emailAddress.strip().lower()
    secretary.facultyName = facultyName.strip()
    secretary.departmentName = departmentName.strip()

    db.commit()
    return {"status": "success"}

# ==========================================
# ADMINISTRATIVE & SECRETARIAT OPERATIONS
# ==========================================

@app.post("/admin/secretariat/delete")
async def delete_secretary(
    id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Deletes a secretariat user account by ID."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    secretary = db.query(Secretariat).filter_by(id=id).first()
    if not secretary:
        raise HTTPException(status_code=404, detail="Secretariat record does not exist")

    db.delete(secretary)
    db.commit()
    return RedirectResponse("/admin/secretariat", status_code=303)


# ==========================================
# ROOM MANAGEMENT & INTEGRATION ENDPOINTS
# ==========================================

@app.get("/secretariat/download-Room-status")
async def download_and_convert_Room(current_user: dict = Depends(get_current_user)):
    """Fetches room data from remote API, saves to Excel, and converts to CSV."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    url = "https://orar.usv.ro/orar/vizualizare/data/sali.php?json"
    folder = "vizualizare_date"
    output_folder = "csv_output"
    os.makedirs(folder, exist_ok=True)
    os.makedirs(output_folder, exist_ok=True)

    excel_path = os.path.join(folder, "Room.xlsx")
    csv_path = os.path.join(output_folder, "Room.csv")

    try:
        response = http_requests.get(url)
        response.raise_for_status()
        data = response.json()

        df = pd.DataFrame(data)
        df = df[df['name'].notna() & (df['name'].str.strip() != "")]
        df = df.sort_values(by='name')

        df.to_excel(excel_path, index=False)
        df.to_csv(csv_path, index=False)

        return {
            "status": "success",
            "messages": [
                "📥 The Room.xlsx file was downloaded successfully.",
                "🔄 The file was converted to Room.csv."
            ],
            "download_url": "/secretariat/download-Room-excel"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading or converting room data: {str(e)}")


@app.get("/secretariat/download-Room-excel")
async def download_Room_excel_file():
    """Serves the generated Excel file containing room data."""
    file_path = os.path.join("vizualizare_date", "Room.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        file_path, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        filename="Room.xlsx"
    )


@app.get("/secretariat/discover-rooms")
async def discover_and_download_Room(
    current_user: dict = Depends(get_current_user)
):
    """Directly downloads and returns the room dataset as an Excel spreadsheet."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    url = "https://orar.usv.ro/orar/vizualizare/data/sali.php?json"
    base_dir = os.path.dirname(__file__)
    folder = os.path.join(base_dir, "vizualizare_date")
    output_file = os.path.join(folder, "Room.xlsx")

    try:
        response = http_requests.get(url)
        response.raise_for_status()
        data = response.json()

        df = pd.DataFrame(data)
        df = df[df['name'].notna() & (df['name'].str.strip() != "")]
        df = df.sort_values(by='name')

        os.makedirs(folder, exist_ok=True)
        df.to_excel(output_file, index=False)

        return FileResponse(
            output_file, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="Room.xlsx"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")


@app.post("/secretariat/upload-rooms")
async def upload_Room(
    file: UploadFile,
    force: bool = Query(False),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Parses uploaded room file and syncs records to database."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    df = pd.read_excel(file.file)

    try:
        if force:
            db.execute(text("DELETE FROM exam_proposals"))
            db.commit()

        db.execute(text("DELETE FROM rooms"))

        cols = ["id", "name", "shortName", "buildingName"]
        col_list = ", ".join([f'"{c}"' for c in cols])
        placeholders = ", ".join([f":{c}" for c in cols])

        for row in df.to_dict(orient="records"):
            db.execute(
                text(f'INSERT INTO rooms({col_list}) VALUES ({placeholders})'),
                row
            )
        db.commit()

        return JSONResponse(status_code=200, content={"imported": len(df)})

    except IntegrityError as e:
        db.rollback()

        if "exam_proposals_room_id_fkey" in str(e.orig):
            raise HTTPException(
                status_code=409,
                detail="ForeignKeyViolation"
            )

        raise HTTPException(
            status_code=500,
            detail=f"Database integrity error: {str(e.orig)}"
        )

    except HTTPException as e:
        return JSONResponse(
            status_code=e.status_code,
            content={"detail": e.detail},
            headers={
                "Access-Control-Allow-Origin": "http://localhost:3000",
                "Access-Control-Allow-Credentials": "true",
            }
        )


@app.options("/secretariat/upload-rooms")
async def options_upload_Room():
    """CORS Preflight handling for room uploads."""
    return JSONResponse(
        status_code=204,
        headers={
            "Access-Control-Allow-Origin": "http://localhost:3000",
            "Access-Control-Allow-Methods": "POST,OPTIONS",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Allow-Credentials": "true",
        },
    )


@app.get("/secretariat", response_class=HTMLResponse)
async def secretariat_dashboard(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Main dashboard view for secretariat users."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    return templates.TemplateResponse("secretariat_Room.html", {"request": request})


@app.post("/secretariat/import-Room-csv")
async def import_Room_csv(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """Imports room definitions directly from cached CSV file using upserts."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    csv_path = "csv_output/Room.csv"
    if not os.path.exists(csv_path):
        raise HTTPException(status_code=404, detail="File Room.csv does not exist")

    try:
        df = pd.read_csv(csv_path)
        count = 0

        for _, row in df.iterrows():
            stmt = pg_insert(Room.__table__).values(
                id=int(row["id"]),
                name=str(row.get("name", "")).strip(),
                shortName=str(row.get("shortName", "")).strip(),
                buildingName=str(row.get("buildingName", "")).strip()
            ).on_conflict_do_update(
                index_elements=["id"],
                set_={
                    "name": str(row.get("name", "")).strip(),
                    "shortName": str(row.get("shortName", "")).strip(),
                    "buildingName": str(row.get("buildingName", "")).strip()
                }
            )
            db.execute(stmt)
            count += 1

        db.commit()
        return {"status": "success", "imported": count}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"CSV import error: {str(e)}")


# ==========================================
# GROUP LEADER MANAGEMENT WORKFLOWS
# ==========================================

@app.get("/secretariat/group-leaders", response_class=HTMLResponse)
async def display_group_leaders(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Renders group leader management overview."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leaders = db.query(GroupLeader).order_by(GroupLeader.lastName, GroupLeader.firstName).all()
    subgroup = db.query(Subgroup).all()
    faculties = db.query(Faculty).all()

    return templates.TemplateResponse("secretariat_group_leaders.html", {
        "request": request,
        "leaders": leaders,
        "Subgroup": subgroup,
        "Faculty": faculties
    })


@app.post("/secretariat/group-leaders/add")
async def add_group_leader(
    request: Request,
    lastName: str = Form(...),
    firstName: str = Form(...),
    emailAddress: str = Form(...),
    phoneNumber: str = Form(None),
    faculty_id: int = Form(...),
    subgroup_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Adds a new group leader entity."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    if db.query(GroupLeader).filter_by(emailAddress=emailAddress.strip().lower()).first():
        return templates.TemplateResponse("secretariat_group_leaders.html", {
            "request": request,
            "leaders": db.query(GroupLeader).all(),
            "Faculty": db.query(Faculty).all(),
            "Subgroup": db.query(Subgroup).all(),
            "error_message": "⚠️ Email address already exists. Please choose another."
        })

    new_leader = GroupLeader(
        lastName=lastName.strip(),
        firstName=firstName.strip(),
        emailAddress=emailAddress.strip().lower(),
        phoneNumber=phoneNumber.strip() if phoneNumber else None,
        faculty_id=faculty_id,
        subgroup_id=subgroup_id
    )

    db.add(new_leader)
    db.commit()

    return RedirectResponse("/secretariat/group-leaders", status_code=303)


@app.post("/secretariat/group-leaders/update")
async def update_group_leaders(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Updates existing group leader attributes."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    form = await request.form()
    try:
        leader_id = int(form.get("id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Invalid Leader ID")

    leader = db.query(GroupLeader).filter_by(id=leader_id).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group leader not found")

    leader.firstName = form.get(f"firstName_{leader_id}", leader.firstName).strip()
    leader.lastName = form.get(f"lastName_{leader_id}", leader.lastName).strip()
    leader.emailAddress = form.get(f"emailAddress_{leader_id}", leader.emailAddress).strip().lower()
    leader.phoneNumber = form.get(f"phoneNumber_{leader_id}", leader.phoneNumber).strip()

    try:
        leader.faculty_id = int(form.get(f"facultyId_{leader_id}"))
        leader.subgroup_id = int(form.get(f"subgroupId_{leader_id}"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Invalid faculty or subgroup ID")

    db.commit()
    return RedirectResponse("/secretariat/group-leaders?msg=saved", status_code=303)


@app.post("/secretariat/group-leaders/delete/{id}")
async def delete_group_leader(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Deletes group leader along with dependent exam proposals."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter_by(id=id).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group leader not found")

    db.query(ExamProposal).filter_by(group_leader_id=id).delete()
    db.delete(leader)
    db.commit()

    return RedirectResponse("/secretariat/group-leaders", status_code=303)


@app.post("/secretariat/group-leaders/import")
async def import_group_leaders(
    excel_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Batch imports group leaders from uploaded Excel workbook."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    try:
        contents = await excel_file.read()
        df = pd.read_excel(io.BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupt Excel spreadsheet")

    try:
        output_csv = os.path.join("csv_output", "group_leaders_import.csv")
        df.to_csv(output_csv, index=False)

        for _, row in df.iterrows():
            if db.query(GroupLeader).filter_by(emailAddress=row["emailAddress"]).first():
                continue

            leader = GroupLeader(
                lastName=row["lastName"],
                firstName=row["firstName"],
                emailAddress=row["emailAddress"],
                phoneNumber=row["phoneNumber"],
                faculty_id=row["faculty_id"],
                subgroup_id=row["subgroup_id"]
            )
            db.add(leader)

        db.commit()
        return {"message": f"{len(df)} group leaders successfully imported."}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")


# ==========================================
# SubjectGENERATION & SCHEDULING CONFIG
# ==========================================

@app.post("/secretariat/subjects/generate")
async def generate_Subjects_endpoint(
    current_user: dict = Depends(get_current_user)
):
    """Queries external schedule API to automatically generate course Subjectassociations."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    excel_path = "vizualizare_date/Subgroup.xlsx"
    df_ids = pd.read_excel(excel_path)
    subgroup_ids = df_ids["id"].dropna().astype(int).tolist()
    data = []

    async with aiohttp.ClientSession() as session:
        for subgroup_id in subgroup_ids:
            try:
                url = f"https://orar.usv.ro/orar/vizualizare/data/orarSPG.php?ID={subgroup_id}&mod=grupa&json"
                json_data = await asyncio.wait_for(fetch_json_with_timeout(session, url), timeout=10)

                if isinstance(json_data, list) and len(json_data) > 0 and isinstance(json_data[0], list):
                    for cls in json_data[0]:
                        if isinstance(cls, dict) and cls.get("typeShortName") == "curs":
                            data.append({
                                "professor_id": int(cls.get("professorID", 0)),
                                "subgroup_id": subgroup_id,
                                "topic": cls.get("topicLongName", "")
                            })

            except Exception:
                continue

    df = pd.DataFrame(data)
    df_unique = df.drop_duplicates(subset=["professor_id", "subgroup_id", "topic"]).reset_index(drop=True)

    with engine.connect() as conn:
        result = conn.execute(text("SELECT id FROM professors"))
        existing_faculty = {row[0] for row in result.fetchall()}

    df_filtered = df_unique[df_unique["professor_id"].isin(existing_faculty)].reset_index(drop=True)
    df_filtered["id"] = df_filtered.index + 1

    os.makedirs("vizualizare_date", exist_ok=True)
    os.makedirs("csv_output", exist_ok=True)

    df_filtered.to_excel("vizualizare_date/Subjects.xlsx", index=False)
    df_filtered.to_csv("csv_output/Subjects.csv", index=False)

    return {"message": f"{len(df_filtered)} Subjects generated", "csv": "csv_output/Subjects.csv"}


@app.post("/secretariat/import-subjects-csv")
async def import_Subjects_csv(
    current_user: dict = Depends(get_current_user)
):
    """Overwrites existing Subjects database table from local CSV dataset."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    csv_path = "csv_output/Subjects.csv"
    if not os.path.exists(csv_path):
        raise HTTPException(status_code=404, detail="File Subjects.csv not found")

    try:
        df = pd.read_csv(csv_path)

        with engine.begin() as conn:
            conn.execute(text("DELETE FROM exam_proposals;"))
            conn.execute(text("DELETE FROM subjects;"))

        df.to_sql("subjects", engine, index=False, if_exists="append")

        return {"message": f"{len(df)} Subjects successfully imported"}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failure: {str(e)}")


@app.get("/secretariat/Subject-admin", response_class=HTMLResponse)
async def admin_Subjects_page(
    request: Request,
    page: int = 1,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Paginated Subjectmanagement view for secretariat."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    per_page = 50
    total = db.query(Subject).count()
    subjects = db.query(Subject).offset((page - 1) * per_page).limit(per_page).all()
    faculty_members = db.query(Professor).all()
    subgroup = db.query(Subgroup).all()

    return templates.TemplateResponse("Subject_admin.html", {
        "request": request,
        "Subjects": subjects,
        "faculty_members": faculty_members,
        "Subgroup": subgroup,
        "current_page": page,
        "total_pages": (total // per_page) + (1 if total % per_page else 0)
    })


@app.post("/secretariat/Subjects/add")
async def add_Subject(
    topic: str = Form(...),
    professor_id: int = Form(...),
    subgroup_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Creates a new course Subjectentry."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    new_Subject= Subject(topic=topic, professor_id=professor_id, subgroup_id=subgroup_id)
    db.add(new_Subject)
    db.commit()
    return RedirectResponse("/secretariat/Subject-admin", status_code=303)


@app.post("/secretariat/Subjects/update/{id}")
async def update_Subject(
    id: int,
    topic: str = Form(...),
    professor_id: int = Form(...),
    subgroup_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Updates an existing Subjectrecord."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    d = db.query(Subject).filter_by(id=id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Subjectnot found")

    d.topic = topic
    d.professor_id = professor_id
    d.subgroup_id = subgroup_id
    db.commit()
    return RedirectResponse("/secretariat/Subject-admin", status_code=303)


@app.get("/secretariat/Subjects/delete/{id}")
async def delete_Subject(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Removes a Subjectby ID."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    subject= db.query(Subject).filter_by(id=id).first()
    if subject:
        db.delete(subject)
        db.commit()
    return RedirectResponse("/secretariat/Subject-admin", status_code=303)


@app.get("/secretariat/exam-limits", response_class=HTMLResponse)
async def exam_limits_form(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Renders form to set examination period start and end bounds."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    limits = db.query(ExamLimits).first()
    return templates.TemplateResponse("exam_limits.html", {
        "request": request,
        "limits": limits
    })


@app.post("/secretariat/exam-limits")
async def set_exam_limits(
    start_date: str = Form(...),
    end_date: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Persists global exam date range constraints."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    start_dt = datetime.fromisoformat(start_date)
    end_dt = datetime.fromisoformat(end_date)

    if start_dt >= end_dt:
        raise HTTPException(status_code=400, detail="Start date must precede end date.")

    limits = db.query(ExamLimits).first()
    if limits:
        limits.start_date = start_dt
        limits.end_date = end_dt
    else:
        limits = ExamLimits(start_date=start_dt, end_date=end_dt)
        db.add(limits)

    db.commit()
    return {
        "message": "Exam limits updated successfully",
        "start_date": limits.start_date.isoformat(),
        "end_date": limits.end_date.isoformat()
    }


@app.get("/secretariat/exam-status")
def get_exam_status(
    request: Request,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=500)
):
    """Returns paginated scheduling status summary across all course Subjects."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    Subjects = db.execute(text("""
        SELECT id, professor_id, subgroup_id, topic
        FROM subjects
        ORDER BY id
    """)).fetchall()

    accepted_exams = db.execute(text("""
        SELECT subject_id
        FROM exam_proposals
        WHERE status = 'accepted'
    """)).fetchall()
    scheduled_Subjects = {row[0] for row in accepted_exams}

    results = []
    scheduled_count = 0

    for d in Subjects:
        is_scheduled = d[0] in scheduled_Subjects
        if is_scheduled:
            scheduled_count += 1

        results.append({
            "id": d[0],
            "professor_id": d[1],
            "subgroup_id": d[2],
            "Subject": d[3],
            "status": "scheduled" if is_scheduled else "unscheduled"
        })

    total = len(results)
    start = (page - 1) * size
    end = start + size
    paginated = results[start:end]
    total_pages = (total + size - 1) // size

    return {
        "total": total,
        "scheduled": scheduled_count,
        "page": page,
        "per_page": size,
        "total_pages": total_pages,
        "Subjects": paginated
    }

# ==========================================
# EXAM SCHEDULING & MANAGEMENT ENDPOINTS
# ==========================================

@app.get("/secretariat/status")
def exam_status_page(
    request: Request, 
    current_user: dict = Depends(get_current_user)
):
    """Renders the exam status overview dashboard for secretariat."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")
    return templates.TemplateResponse("secretariat_exam_status.html", {"request": request})


@app.get("/secretariat/exams", response_class=HTMLResponse)
def list_scheduled_exams(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Renders HTML view listing all accepted/scheduled exams."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    exams = db.query(ExamProposal).filter(
        ExamProposal.status == "accepted"
    ).options(
        joinedload(ExamProposal.subject),
        joinedload(ExamProposal.room),
        joinedload(ExamProposal.assistant)
    ).all()

    return templates.TemplateResponse("secretariat_exams_list.html", {
        "request": request,
        "exams": exams
    })


@app.get("/api/scheduled-exams")
def get_scheduled_exams(
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """API endpoint returning array of accepted exam schedule records."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    exams = db.execute(text("""
        SELECT id, subject_id, room_id, date, duration, assistant_id
        FROM exam_proposals
        WHERE status = 'accepted'
        ORDER BY date
    """)).fetchall()

    return [dict(row._mapping) for row in exams]


@app.delete("/api/scheduled-exams/{exam_id}")
def delete_scheduled_exam(
    exam_id: int, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Removes an accepted exam entry by ID."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    db.execute(text("DELETE FROM exam_proposals WHERE id = :id AND status = 'accepted'"), {"id": exam_id})
    db.commit()
    return {"message": "Exam entry successfully deleted"}


@app.put("/api/scheduled-exams/{exam_id}")
def update_scheduled_exam(
    exam_id: int, 
    payload: dict, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Updates timing, room, and assistant parameters for a scheduled exam."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    db.execute(text("""
        UPDATE exam_proposals
        SET room_id = :room_id,
            date = :date,
            duration = :duration,
            assistant_id = :assistant_id
        WHERE id = :id AND status = 'accepted'
    """), {
        "id": exam_id,
        "room_id": payload.get("room_id"),
        "date": payload.get("date"),
        "duration": payload.get("duration"),
        "assistant_id": payload.get("assistant_id")
    })
    db.commit()
    return {"message": "Exam details successfully updated"}


@app.post("/api/scheduled-exams")
def create_scheduled_exam(
    payload: dict, 
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Directly inserts a pre-approved exam record."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    db.execute(text("""
        INSERT INTO exam_proposals
        (
            subject_id,
            group_leader_id,
            room_id,
            date,
            duration,
            assistant_id,
            status
        )
        VALUES
        (
            :subject_id,
            :group_leader_id,
            :room_id,
            :date,
            :duration,
            :assistant_id,
            'accepted'
        )
    """), payload)
    db.commit()
    return {"message": "Exam successfully created"}


@app.get("/secretariat/exam/{id}", response_class=HTMLResponse)
def edit_exam_form(
    id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Renders edit interface for an individual exam proposal."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    exam = db.query(ExamProposal).filter_by(id=id, status="accepted").first()
    if not exam:
        raise HTTPException(status_code=404, detail="Exam record not found")

    # Fetch available Room(excluding Roombooked at the same date)
    all_Room= db.query(Room).all()
    occupied_Room= db.query(ExamProposal.room_id).filter(
        ExamProposal.date == exam.date,
        ExamProposal.id != id,
        ExamProposal.status == "accepted"
    ).all()
    blocked_room_ids = {r.room_id for r in occupied_Room}
    available_Room= [r for r in all_Room if r.id not in blocked_room_ids]

    # Department teaching staff for assistant selection
    faculty_staff = db.query(Professor).filter(
        Professor.departmentName == exam.subject.professor.departmentName
    ).all()

    return templates.TemplateResponse("secretariat_exam_edit.html", {
        "request": request,
        "exam": exam,
        "Room": available_Room,
        "assistants": faculty_staff
    })


@app.post("/secretariat/exam/{id}", response_class=HTMLResponse)
def save_edited_exam(
    id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    date: str = Form(...),
    duration: int = Form(...),
    room_id: int = Form(...),
    assistant_id: Optional[int] = Form(None)
):
    """Validates room/assistant schedule overlaps and saves exam revisions."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    exam = db.query(ExamProposal).filter_by(id=id, status="accepted").first()
    if not exam:
        return templates.TemplateResponse("secretariat_exam_edit.html", {
            "request": request,
            "exam": None,
            "Room": [],
            "assistants": [],
            "error_message": "Exam record not found"
        }, status_code=404)

    new_start = datetime.fromisoformat(date)
    new_end = new_start + timedelta(hours=duration)

    # Check for room overlap
    room_conflicts = db.query(ExamProposal).filter(
        ExamProposal.id != id,
        ExamProposal.status == "accepted",
        ExamProposal.room_id == room_id
    ).all()

    for e in room_conflicts:
        e_start = e.date
        e_end = e_start + timedelta(hours=e.duration)
        if new_start < e_end and new_end > e_start:
            all_Room= db.query(Room).all()
            occupied_Room= db.query(ExamProposal.room_id).filter(
                ExamProposal.date == exam.date,
                ExamProposal.id != id,
                ExamProposal.status == "accepted"
            ).all()
            blocked_room_ids = {r.room_id for r in occupied_Room}
            available_Room= [r for r in all_Room if r.id not in blocked_room_ids]
            faculty_staff = db.query(Professor).filter(
                Professor.departmentName == exam.subject.professor.departmentName
            ).all()

            return templates.TemplateResponse("secretariat_exam_edit.html", {
                "request": request,
                "exam": exam,
                "Room": available_Room,
                "assistants": faculty_staff,
                "error_message": "Selected room is already occupied during this time window."
            }, status_code=200)

    # Check for assistant overlap
    if assistant_id:
        assistant_conflicts = db.query(ExamProposal).join(Subject).filter(
            ExamProposal.id != id,
            ExamProposal.status == "accepted",
            or_(
                ExamProposal.assistant_id == assistant_id,
                Subject.professor_id == assistant_id
            )
        ).all()

        for p in assistant_conflicts:
            p_start = p.date
            p_end = p_start + timedelta(hours=p.duration)
            if new_start < p_end and new_end > p_start:
                all_Room= db.query(Room).all()
                occupied_Room= db.query(ExamProposal.room_id).filter(
                    ExamProposal.date == exam.date,
                    ExamProposal.id != id,
                    ExamProposal.status == "accepted"
                ).all()
                blocked_room_ids = {r.room_id for r in occupied_Room}
                available_Room= [r for r in all_Room if r.id not in blocked_room_ids]
                faculty_staff = db.query(Professor).filter(
                    Professor.departmentName == exam.subject.professor.departmentName
                ).all()

                return templates.TemplateResponse("secretariat_exam_edit.html", {
                    "request": request,
                    "exam": exam,
                    "Room": available_Room,
                    "assistants": faculty_staff,
                    "error_message": "Selected assistant is already assigned to an exam during this time window."
                }, status_code=200)

    # Persist changes
    exam.date = new_start
    exam.duration = duration
    exam.room_id = room_id
    exam.assistant_id = assistant_id if assistant_id else None

    db.commit()
    return RedirectResponse(url="/secretariat/exams", status_code=303)


@app.get("/secretariat/available-Room")
def get_available_Room_endpoint(
    date: str,
    duration: int = 90,
    exam_id: int = 0,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Calculates room availability given target date, duration, and optional excluded exam ID."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    try:
        start = datetime.fromisoformat(date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    end = start + timedelta(hours=duration)

    conflicting_exams = db.query(ExamProposal).filter(
        ExamProposal.status == "accepted",
        ExamProposal.id != exam_id
    ).all()

    occupied_room_ids = set()
    for ex in conflicting_exams:
        ex_start = ex.date
        ex_end = ex_start + timedelta(hours=ex.duration)
        if start < ex_end and end > ex_start:
            occupied_room_ids.add(ex.room_id)

    available_Room= db.query(Room).filter(~Room.id.in_(occupied_room_ids)).all()

    return [{"id": r.id, "name": r.name} for r in available_Room]


# ==========================================
# REPORTING & EXPORT ENDPOINTS (EXCEL / PDF)
# ==========================================

@app.get("/secretariat/schedule/excel")
def export_schedule_excel(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Exports scheduled examination calendar to downloadable Excel file."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    exams = db.query(ExamProposal).filter(
        ExamProposal.status == "accepted"
    ).options(
        joinedload(ExamProposal.subject).joinedload(Subject.professor),
        joinedload(ExamProposal.room),
        joinedload(ExamProposal.assistant)
    ).all()

    # Sort entries by subgroup code
    exams.sort(
        key=lambda e: f"{e.subject.subgroup.groupName}{e.subject.subgroup.subgroupIndex}"
    )

    rows = []
    for e in exams:
        subgroup_code = f"{e.subject.subgroup.groupName}{e.subject.subgroup.subgroupIndex}"
        study_year = e.subject.subgroup.studyYear or "—"

        rows.append({
            "Subject": e.subject.topic,
            "Course Leader": f"{e.subject.professor.firstName} {e.subject.professor.lastName}",
            "Exam Date": e.date.strftime("%Y-%m-%d %H:%M"),
            "Duration (h)": e.duration,
            "Room": e.room.name if e.room else "—",
            "Assistant": f"{e.assistant.firstName} {e.assistant.lastName}" if e.assistant else "—",
            "Subgroup": subgroup_code,
            "Year": study_year
        })

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Exam Schedule")

    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=exam_schedule.xlsx"}
    )


@app.get("/secretariat/schedule/pdf")
def export_schedule_pdf(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Exports scheduled examination calendar to formatted landscape PDF document."""
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access forbidden")

    exams = db.query(ExamProposal).filter(
        ExamProposal.status == "accepted"
    ).options(
        joinedload(ExamProposal.subject).joinedload(Subject.professor),
        joinedload(ExamProposal.room),
        joinedload(ExamProposal.assistant)
    ).all()

    exams.sort(
        key=lambda e: f"{e.subject.subgroup.groupName}{e.subject.subgroup.subgroupIndex}"
    )

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.add_font("DejaVu", "", "static/fonts/DejaVuSans.ttf", uni=True)
    pdf.add_font("DejaVu", "B", "static/fonts/DejaVuSans-Bold.ttf", uni=True)
    pdf.set_font("DejaVu", size=10)

    pdf.set_fill_color(200, 220, 255)
    pdf.cell(0, 10, "Exam Schedule", ln=True, align="C")

    # Table headers
    headers = ["Subject", "Leader", "Date", "Duration", "Room", "Assistant", "Subgroup", "Year"]
    col_widths = [50, 40, 30, 18, 25, 40, 30, 15]
    pdf.set_font("DejaVu", size=9)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, 1, 0, "C", fill=True)
    pdf.ln()

    last_subgroup = None
    for e in exams:
        subgroup_code = f"{e.subject.subgroup.groupName}{e.subject.subgroup.subgroupIndex}"
        study_year = str(e.subject.subgroup.studyYear) if e.subject.subgroup.studyYear else "—"

        # Subgroup section divider header
        if subgroup_code != last_subgroup:
            pdf.set_font("DejaVu", style="B", size=10)
            pdf.set_fill_color(230, 230, 250)
            pdf.cell(sum(col_widths), 8, f"Subgroup: {subgroup_code} | Year: {study_year}", 1, ln=1, fill=True)
            pdf.set_font("DejaVu", size=9)
            last_subgroup = subgroup_code

        values = [
            e.subject.topic,
            f"{e.subject.professor.firstName} {e.subject.professor.lastName}",
            e.date.strftime("%Y-%m-%d %H:%M"),
            f"{e.duration} h",
            e.room.name if e.room else "—",
            f"{e.assistant.firstName} {e.assistant.lastName}" if e.assistant else "—",
            subgroup_code,
            study_year
        ]

        for i, val in enumerate(values):
            pdf.cell(col_widths[i], 10, str(val), 1, 0, "L")
        pdf.ln()

    pdf_bytes = pdf.output(dest="S").encode("latin1")

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=exam_schedule.pdf"}
    )


# ==========================================
# professor / FACULTY MEMBER PROPOSALS API
# ==========================================

@app.get("/professor/proposals/json")
async def get_professor_proposals_json(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Returns submitted proposals and room/assistant availability lists for active faculty member."""
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Access restricted to teaching staff")

    professor = db.query(Professor).filter(Professor.emailAddress == current_user["email"]).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    proposals = db.query(ExamProposal).join(Subject).filter(
        Subject.professor_id == professor.id,
        ExamProposal.status == "submitted"
    ).all()

    all_rooms= db.query(Room).all()
    colleagues = db.query(Professor).filter(
        Professor.departmentName == professor.departmentName,
        Professor.id != professor.id
    ).all()

    available_rooms= {}
    available_assistants = {}

    for prop in proposals:
        start = prop.date.replace(tzinfo=None)
        end = start + timedelta(hours=prop.duration)

        accepted_exams = db.query(ExamProposal).filter(
            ExamProposal.room_id.isnot(None),
            ExamProposal.status == "accepted"
        ).all()

        occupied_room_ids = {
            p.room_id for p in accepted_exams
            if start < p.date.replace(tzinfo=None) + timedelta(hours=p.duration)
            and end > p.date.replace(tzinfo=None)
        }

        available_rooms[prop.id] = [
            {"id": r.id, "name": r.name}
            for r in all_rooms if r.id not in occupied_room_ids
        ]

        available_list = []
        for colleague in colleagues:
            colleague_exams = db.query(ExamProposal).filter(
                ExamProposal.assistant_id == colleague.id,
                ExamProposal.status == "accepted"
            ).all()

            has_conflict = any(
                start < ep.date.replace(tzinfo=None) + timedelta(hours=ep.duration) and
                end > ep.date.replace(tzinfo=None)
                for ep in colleague_exams
            )
            if not has_conflict:
                available_list.append({
                    "id": colleague.id,
                    "firstName": colleague.firstName,
                    "lastName": colleague.lastName
                })

        available_assistants[prop.id] = available_list

    return {
        "professor": {
            "firstName": professor.firstName,
            "lastName": professor.lastName
        },
        "proposals": [
            {
                "id": p.id,
                "subject": {
                    "topic": p.subject.topic,
                    "subgroup": {
                        "studyYear": p.subject.subgroup.studyYear,
                        "groupName": p.subject.subgroup.groupName,
                        "subgroupIndex": p.subject.subgroup.subgroupIndex
                    }
                },
                "date": p.date.isoformat(),
                "duration": p.duration
            }
            for p in proposals
        ],
        "available_assistants": available_assistants,
        "available_rooms": available_rooms
    }


# ==========================================
# USER PROFILES & ADMIN API
# ==========================================

@app.get("/me")
async def get_current_user_profile(
    current_user: dict = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    """Unified user context endpoint returning profile fields based on user role."""
    user = None
    if current_user["role"] == "admin":
        user = db.query(Admin).filter_by(emailAddress=current_user["email"]).first()
    elif current_user["role"] == "professor":
        user = db.query(Professor).filter_by(emailAddress=current_user["email"]).first()
    elif current_user["role"] == "secretariat":
        user = db.query(Secretariat).filter_by(emailAddress=current_user["email"]).first()
    elif current_user["role"] == "group_leader":
        user = db.query(GroupLeader).filter_by(emailAddress=current_user["email"]).first()

    if not user:
        raise HTTPException(status_code=404, detail="User profile not found")

    return {
        "email": current_user["email"],
        "firstName": user.firstName,
        "lastName": user.lastName,
        "phoneNumber": user.phoneNumber,
        "role": current_user["role"]
    }


@app.get("/api/admin/faculties")
async def list_faculties_json(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    faculties = db.query(Faculty).order_by(Faculty.id.desc()).all()

    return [
        {
            "id": faculty.id,
            "longName": faculty.longName,
            "shortName": faculty.shortName
        }
        for faculty in faculties
    ]


@app.get("/api/admin/faculty-members")
async def admin_faculty_members_json(
    db: Session = Depends(get_db), 
    current_user: dict = Depends(get_current_user)
):
    """Returns faculty members with university email domain."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    faculty_list = db.query(Professor).filter(Professor.emailAddress.ilike('%@usm.ro')).all()
    return [
        {
            "id": c.id,
            "firstName": c.firstName,
            "lastName": c.lastName,
            "emailAddress": c.emailAddress,
            "phoneNumber": c.phoneNumber,
            "facultyName": c.facultyName,
            "departmentName": c.departmentName,
        } for c in faculty_list
    ]



@app.get("/admin/secretariat/json")
async def list_secretaries_json(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Returns list of secretariat staff members."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    secretaries = db.query(Secretariat).order_by(Secretariat.lastName, Secretariat.firstName).all()
    return [
        {
            "id": s.id,
            "firstName": s.firstName,
            "lastName": s.lastName,
            "emailAddress": s.emailAddress,
            "facultyName": s.facultyName,
            "departmentName": s.departmentName
        } for s in secretaries
    ]


@app.get("/admin/profile", response_model=dict)
async def get_admin_profile(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Returns profile details for currently logged-in administrator."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    admin = db.query(Admin).filter_by(id=current_user["id"]).first()
    if not admin:
        raise HTTPException(status_code=404, detail="Admin record not found")

    return {
        "id": admin.id,
        "firstName": admin.firstName,
        "lastName": admin.lastName,
        "emailAddress": admin.emailAddress,
        "facultyName": admin.facultyName,
        "departmentName": admin.departmentName
    }


@app.post("/admin/profile/update")
async def update_admin_profile(
    id: int = Form(...),
    firstName: str = Form(...),
    lastName: str = Form(...),
    facultyName: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Updates administrator profile information."""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    admin = db.query(Admin).filter_by(id=id).first()
    if not admin:
        raise HTTPException(status_code=404, detail="Admin record not found")

    admin.firstName = firstName.strip()
    admin.lastName = lastName.strip()
    admin.facultyName = facultyName.strip()
    admin.departmentName = departmentName.strip()

    db.commit()
    return {"status": "success"}

@app.post("/admin/faculties/update")
async def update_faculties(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    form = await request.form()
    faculties = db.query(Faculty).all()

    for faculty in faculties:
        long_name = form.get(f"longName_{faculty.id}")
        short_name = form.get(f"shortName_{faculty.id}")

        if long_name and short_name:
            faculty.longName = long_name.strip()
            faculty.shortName = short_name.strip()

    db.commit()

    return {"message": "Faculties updated successfully"}

@app.post("/admin/faculties/add")
async def add_faculty(
    longName: str = Form(...),
    shortName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    new_faculty = Faculty(
        longName=longName.strip(),
        shortName=shortName.strip()
    )

    db.add(new_faculty)
    db.commit()

    return {"message": "Faculty added successfully"}

@app.post("/admin/faculties/delete")
async def delete_faculty(
    id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Access forbidden")

    faculty = db.query(Faculty).filter(Faculty.id == id).first()

    if not faculty:
        raise HTTPException(status_code=404, detail="Faculty not found")

    subgroup_count = (
        db.query(Subgroup)
        .filter(Subgroup.facultyId == id)
        .count()
    )

    if subgroup_count > 0:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete faculty because related subgroups exist."
        )

    db.delete(faculty)
    db.commit()

    return {"message": "Faculty deleted successfully"}

# ==========================================
# GROUP LEADER DASHBOARD API
# ==========================================

@app.get(
    "/api/group-leader/subject-status",
    response_class=JSONResponse
)
async def group_leader_subject_status_json(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user["role"] != "group_leader":
        raise HTTPException(
            status_code=403,
            detail="Access forbidden"
        )

    leader = db.query(GroupLeader).filter(
        GroupLeader.emailAddress == current_user["email"]
    ).first()

    if not leader:
        raise HTTPException(
            status_code=404,
            detail="Group leader not found"
        )

    subjects = db.query(Subject).filter(
        Subject.subgroup_id == leader.subgroup_id
    ).all()

    proposals = db.query(ExamProposal).filter(
        ExamProposal.group_leader_id == leader.id
    ).all()

    proposal_map = {
        proposal.subject_id: proposal
        for proposal in proposals
    }

    submitted = []
    accepted = []
    rejected = []
    unsubmitted = []

    for subject in subjects:
        proposal = proposal_map.get(subject.id)
        subgroup = subject.subgroup

        info = {
            "id": subject.id,
            "subjectName": subject.topic,
            "year": subgroup.studyYear,
            "group": (
                f"{subgroup.groupName}"
                f"{subgroup.subgroupIndex}"
            )
        }

        if proposal:
            info.update({
                "examDate": proposal.date.isoformat(),
                "duration": proposal.duration,
                "status": proposal.status,
                "rejectionReason": proposal.rejection_reason
            })

            if proposal.status == "submitted":
                submitted.append(info)
            elif proposal.status == "accepted":
                accepted.append(info)
            elif proposal.status == "rejected":
                rejected.append(info)
        else:
            unsubmitted.append(info)

    return {
        "submitted": submitted,
        "accepted": accepted,
        "rejected": rejected,
        "unsubmitted": unsubmitted,
        "groupLeader": {
            "id": leader.id,
            "firstName": leader.firstName,
            "lastName": leader.lastName
        }
    }


@app.get("/api/group-leader/unproposed-subjects")
async def get_unproposed_Subjects(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Lists Subjects assigned to the group leader's subgroup that have no existing proposal."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter(GroupLeader.emailAddress == current_user["email"]).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group leader not found")

    subjects = db.query(Subject).filter(Subject.subgroup_id == leader.subgroup_id).all()
    proposals = db.query(ExamProposal).filter(ExamProposal.group_leader_id == leader.id).all()
    proposed_ids = {p.subject_id for p in proposals}

    return [
        {"id": subject.id, "subjectName": subject.topic}
        for subject in subjects if subject.id not in proposed_ids
    ]


@app.get("/api/group-leader/accepted-exams")
async def get_accepted_exams_group_leader(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Retrieves all accepted exams relevant to the group leader's assigned subgroup."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter(GroupLeader.emailAddress == current_user["email"]).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group leader not found")

    exams = (
        db.query(ExamProposal, Subject)
        .join(Subject, Subject.id == ExamProposal.subject_id)
        .filter(Subject.subgroup_id == leader.subgroup_id)
        .filter(ExamProposal.status == "accepted")
        .all()
    )

    return [
        {
            "id": exam.id,
            "subject": subject.topic,
            "date": exam.date.isoformat(),
            "room": exam.room.name if exam.room else None
        }
        for exam, subject in exams
    ]

@app.put("/group-leader/update")
async def update_group_leader_profile(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Updates group leader profile details."""
    payload = await request.json()

    if current_user["role"] != "group_leader":
        raise HTTPException(
            status_code=403,
            detail="Access restricted to group leaders"
        )

    leader = db.query(GroupLeader).filter_by(id=current_user["id"]).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group leader record not found")

    for key in ["firstName", "lastName", "phoneNumber"]:
        if key in payload:
            setattr(leader, key, payload[key])

    db.add(leader)
    db.commit()
    return {"status": "ok"}


@app.get("/group-leader/me")
async def get_group_leader_me(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Returns current group leader details."""
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access forbidden")

    leader = db.query(GroupLeader).filter_by(id=current_user["id"]).first()
    if not leader:
        raise HTTPException(status_code=404, detail="Group leader record not found")

    return {
        "id": leader.id,
        "email": leader.emailAddress,
        "firstName": leader.firstName,
        "lastName": leader.lastName,
        "phoneNumber": leader.phoneNumber or ""
    }

@app.get("/professor/accepted-exams/json")
def get_accepted_exams_json(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # 1. Check role and professor existence
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Access denied")
    
    professor = db.query(Professor).filter(Professor.emailAddress == current_user["email"]).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    # 2. Gather exam proposals with status="accepted"
    exams = (
        db.query(ExamProposal)
        .join(Subject)
        .filter(
            Subject.professor_id == professor.id,
            ExamProposal.status == "accepted"
        )
        .all()
    )

    # 3. Build JSON response list
    json_list = []
    for e in exams:
        subgroup = e.subject.subgroup
        room_obj = e.room
        assistant_obj = e.assistant

        json_list.append({
            "id": e.id,
            "subject": e.subject.topic,
            "studyYear": subgroup.studyYear if subgroup else None,
            "groupName": subgroup.groupName if subgroup else None,
            "subgroupIndex": subgroup.subgroupIndex if subgroup else None,
            "date": e.date.isoformat(),
            "duration": e.duration,
            "room": room_obj.name if room_obj else None,
            "assistantFirstName": assistant_obj.firstName if assistant_obj else None,
            "assistantLastName": assistant_obj.lastName if assistant_obj else None
        })

    return JSONResponse({
        "professor": {
            "firstName": professor.firstName,
            "lastName": professor.lastName
        },
        "exams": json_list
    })


@app.get("/professor/export-excel-assistant")
async def export_assistant_exams_excel(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # 1) Check role and get Professor object
    if current_user["role"] != "professor":
        raise HTTPException(status_code=403, detail="Access permitted to teaching staff only")

    professor = db.query(Professor).filter_by(emailAddress=current_user["email"]).first()
    if not professor:
        raise HTTPException(status_code=404, detail="Professor not found")

    # 2) Create new workbook
    wb = Workbook()
    assistant_ws = wb.active
    assistant_ws.title = "Assistant Exams"

    # 3) Add column headers
    assistant_ws.append([
        "Subject", 
        "Year", 
        "Group", 
        "Date and Time", 
        "Duration", 
        "Room", 
        "Lead professor"
    ])

    # 4) Fetch proposals where professor acts as assistant
    assistant_exams = (
        db.query(ExamProposal)
        .join(Subject)
        .join(Subgroup)
        .filter(
            ExamProposal.assistant_id == professor.id,
            ExamProposal.status == "accepted"
        )
        .all()
    )

    # 5) Populate worksheet rows
    for p in assistant_exams:
        subgroup = p.subject.subgroup
        group_str = f"{subgroup.studyYear} {subgroup.groupName}{subgroup.subgroupIndex}" if subgroup else "-"
        room_name = p.room.name if p.room else "-"
        lead_professor = f"{p.subject.professor.firstName} {p.subject.professor.lastName}" if (p.subject and p.subject.professor) else "-"

        assistant_ws.append([
            p.subject.topic if p.subject else "-",
            subgroup.studyYear if subgroup else "-",
            group_str,
            p.date.strftime("%Y-%m-%d %H:%M"),
            p.duration,
            room_name,
            lead_professor
        ])

    # 6) Save workbook to BytesIO and return stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=assistant_exams.xlsx"
        }
    )


@app.get("/secretariat/api/rooms")
async def api_list_Room(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")
    rooms = db.query(Room).order_by(Room.id).all()
    return [{"id": r.id, "name": r.name, "shortName": r.shortName, "buildingName": r.buildingName} for r in rooms]


@app.post("/secretariat/api/rooms")
async def api_create_room(
    payload: dict = Body(..., example={"name": "Room 1", "shortName": "R1", "buildingName": "Building A"}),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    max_id = db.query(func.max(Room.id)).scalar() or 0
    next_id = max_id + 1

    room = Room(id=next_id, **payload)
    db.add(room)
    db.commit()
    db.refresh(room)

    return {"id": room.id, **payload}


@app.put("/secretariat/api/rooms/{id}")
async def api_update_room(
    id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")
    room = db.query(Room).get(id)
    if not room:
        raise HTTPException(404, "Room does not exist")
    for k, v in payload.items():
        setattr(room, k, v)
    db.commit()
    db.refresh(room)
    return {
        "id": room.id,
        "name": room.name,
        "shortName": room.shortName,
        "buildingName": room.buildingName
    }


@app.delete("/secretariat/api/rooms/{id}")
async def api_delete_room(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")
    room = db.query(Room).get(id)
    if not room:
        raise HTTPException(404, "Room not found")
    db.delete(room)
    db.commit()
    return JSONResponse(status_code=200, content={"deleted": id})


@app.post("/secretariat/import-group-leaders")
async def import_group_leaders_csv(
    file: UploadFile,
    force: bool = Query(False),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    sec = db.query(Secretariat).get(current_user["id"])
    if not sec:
        raise HTTPException(404, "Secretariat user does not exist")
    facultyName = sec.facultyName

    faculty = db.query(Faculty).filter(Faculty.longName == facultyName).first()
    if not faculty:
        raise HTTPException(404, f"Faculty '{facultyName}' does not exist")
    faculty_id = faculty.id

    max_id = db.query(func.max(GroupLeader.id)).scalar() or 0
    next_id = max_id + 1

    try:
        df = pd.read_excel(file.file)
    except Exception as e:
        raise HTTPException(400, f"Invalid file: {e}")

    required_cols = {"lastName", "firstName", "emailAddress", "subgroup_id", "faculty_id"}
    missing = required_cols - set(df.columns)
    if missing:
        raise HTTPException(400, f"Missing column(s): {', '.join(sorted(missing))}")

    df = df[df["faculty_id"] == faculty_id]

    inserted = 0
    updated = 0

    for raw in df.to_dict(orient="records"):
        sub_id = int(raw["subgroup_id"])
        existing = (
            db.query(GroupLeader)
            .filter_by(faculty_id=faculty_id, subgroup_id=sub_id)
            .first()
        )

        if existing:
            existing.lastName = raw["lastName"].strip()
            existing.firstName = raw["firstName"].strip()
            existing.emailAddress = raw["emailAddress"].strip()
            existing.phoneNumber = raw.get("phoneNumber")
            updated += 1
        else:
            group_leader = GroupLeader(
                id=next_id,
                lastName=raw["lastName"].strip(),
                firstName=raw["firstName"].strip(),
                emailAddress=raw["emailAddress"].strip(),
                phoneNumber=raw.get("phoneNumber"),
                faculty_id=faculty_id,
                subgroup_id=sub_id
            )
            db.add(group_leader)
            next_id += 1
            inserted += 1

    db.commit()

    return JSONResponse(
        status_code=200,
        content={
            "imported": inserted,
            "updated": updated,
            "message": f"✔ New inserts: {inserted}, Updates: {updated} in faculty '{facultyName}'."
        }
    )


@app.get("/secretariat/api/group-leaders")
async def api_list_group_leaders(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    sec = db.query(Secretariat).get(current_user["id"])
    if not sec:
        raise HTTPException(404, "Secretariat user does not exist")
    faculty = db.query(Faculty).filter(Faculty.longName == sec.facultyName).first()
    if not faculty:
        raise HTTPException(404, f"Faculty '{sec.facultyName}' does not exist")
    faculty_id = faculty.id

    group_leaders = (
        db.query(GroupLeader)
        .filter(GroupLeader.faculty_id == faculty_id)
        .order_by(GroupLeader.id)
        .all()
    )
    return [
        {
            "id": s.id,
            "lastName": s.lastName,
            "firstName": s.firstName,
            "emailAddress": s.emailAddress,
            "phoneNumber": s.phoneNumber,
            "subgroup_id": s.subgroup_id,
            "studyYear": s.subgroup.studyYear if s.subgroup else None,
            "groupName": s.subgroup.groupName if s.subgroup else None,
            "subgroupIndex": s.subgroup.subgroupIndex if s.subgroup else None
        }
        for s in group_leaders
    ]


@app.post("/secretariat/api/group-leaders")
async def api_create_group_leader(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    sec = db.query(Secretariat).get(current_user["id"])
    if not sec:
        raise HTTPException(404, "Secretariat user does not exist")
    faculty = db.query(Faculty).filter(Faculty.longName == sec.facultyName).first()
    if not faculty:
        raise HTTPException(404, f"Faculty '{sec.facultyName}' does not exist")
    faculty_id = faculty.id

    group_leader = GroupLeader(
        lastName=payload["lastName"],
        firstName=payload["firstName"],
        emailAddress=payload["emailAddress"],
        phoneNumber=payload.get("phoneNumber"),
        faculty_id=faculty_id,
        subgroup_id=payload["subgroup_id"]
    )
    db.add(group_leader)
    db.commit()
    db.refresh(group_leader)

    return {
        "id": group_leader.id,
        **payload,
        "studyYear": group_leader.subgroup.studyYear if group_leader.subgroup else None,
        "groupName": group_leader.subgroup.groupName if group_leader.subgroup else None,
        "subgroupIndex": group_leader.subgroup.subgroupIndex if group_leader.subgroup else None
    }


@app.put("/secretariat/api/group-leaders/{id}")
async def api_update_group_leader(
    id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    group_leader = db.query(GroupLeader).get(id)
    if not group_leader:
        raise HTTPException(404, "Group leader does not exist")

    for k in ["lastName", "firstName", "emailAddress", "phoneNumber", "subgroup_id"]:
        if k in payload:
            setattr(group_leader, k, payload[k])
    db.commit()
    db.refresh(group_leader)

    return {
        "id": group_leader.id,
        "lastName": group_leader.lastName,
        "firstName": group_leader.firstName,
        "emailAddress": group_leader.emailAddress,
        "phoneNumber": group_leader.phoneNumber,
        "subgroup_id": group_leader.subgroup_id,
        "studyYear": group_leader.subgroup.studyYear if group_leader.subgroup else None,
        "groupName": group_leader.subgroup.groupName if group_leader.subgroup else None,
        "subgroupIndex": group_leader.subgroup.subgroupIndex if group_leader.subgroup else None
    }


@app.delete("/secretariat/api/group-leaders/{id}")
async def api_delete_group_leader(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    group_leader = db.query(GroupLeader).get(id)
    if not group_leader:
        raise HTTPException(404, "Group leader does not exist")
    db.query(ExamProposal).filter(
        ExamProposal.group_leader_id == id
    ).delete(synchronize_session=False)
    db.delete(group_leader)
    db.commit()
    return JSONResponse(status_code=200, content={"deleted": id})


@app.get("/secretariat/api/subgroups")
async def api_list_Subgroup(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    sec = db.query(Secretariat).get(current_user["id"])
    if not sec:
        raise HTTPException(404, "Secretariat user does not exist")
    facultyName = sec.facultyName

    faculty = db.query(Faculty).filter(Faculty.longName == facultyName).first()
    if not faculty:
        raise HTTPException(404, f"Faculty '{facultyName}' does not exist")
    faculty_id = faculty.id

    subgroups = (
        db.query(Subgroup)
        .filter(Subgroup.facultyId == faculty_id)
        .order_by(Subgroup.studyYear, Subgroup.groupName, Subgroup.subgroupIndex)
        .all()
    )

    return [
        {
            "id": sg.id,
            "studyYear": sg.studyYear,
            "groupName": sg.groupName,
            "subgroupIndex": sg.subgroupIndex
        }
        for sg in subgroups
    ]


@app.get(
    "/secretariat/api/professors",
    response_model=list[dict]
)
def list_secretariat_professors(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )

    secretariat = db.query(Secretariat).filter_by(
        id=current_user["id"]
    ).first()

    if not secretariat:
        raise HTTPException(
            status_code=404,
            detail="Secretariat not found"
        )

    professors = (
        db.query(Professor)
        .filter(
            Professor.facultyName ==
            secretariat.facultyName
        )
        .order_by(
            Professor.lastName,
            Professor.firstName
        )
        .all()
    )

    return [
        {
            "id": professor.id,
            "firstName": professor.firstName,
            "lastName": professor.lastName,
            "departmentName": professor.departmentName
        }
        for professor in professors
    ]

@app.get("/secretariat/subjects", response_class=HTMLResponse)
async def subjects_admin_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access denied")

    sec = db.query(Secretariat).filter_by(id=current_user["id"]).first()
    faculty = db.query(Faculty).filter_by(longName=sec.facultyName).first()

    subgroup_ids = [s.id for s in db.query(Subgroup).filter_by(facultyId=faculty.id).all()]
    subjects = db.query(Subject).filter(Subject.subgroup_id.in_(subgroup_ids)).all()

    return templates.TemplateResponse("secretariat_subjects_admin.html", {
        "request": request,
        "subjects": subjects
    })

@app.get("/secretariat/api/subjects")
async def api_list_subjects(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 1) Authorization
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    # 2) Find the secretariat's faculty
    sec = db.query(Secretariat).get(current_user["id"])
    if not sec:
        raise HTTPException(404, "Secretariat user does not exist")
    fac = (
        db.query(Faculty)
          .filter(Faculty.longName == sec.facultyName)
          .first()
    )
    if not fac:
        raise HTTPException(404, f"Faculty '{sec.facultyName}' does not exist")
    fac_id = fac.id

    # 3) List only subjects for Subgroup belonging to the faculty
    sub_ids = [sg.id for sg in fac.subgroups]
    subjects = (
        db.query(Subject)
          .filter(Subject.subgroup_id.in_(sub_ids))
          .order_by(Subject.id)
          .all()
    )

    # 4) Return the list
    return [
        {
            "id":          s.id,
            "topic":       s.topic,
            "id_subgroup": s.subgroup_id,
            "professor_id":  s.professor_id
        }
        for s in subjects
    ]


# ──────────────────────────────────────────────────────────────────────────────
@app.post("/secretariat/api/subjects")
async def api_create_subject(
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # Authorization
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    # Determine faculty and Subgroup
    sec = db.query(Secretariat).get(current_user["id"])
    if not sec:
        raise HTTPException(404, "Secretariat user does not exist")
    fac = (
        db.query(Faculty)
          .filter(Faculty.longName == sec.facultyName)
          .first()
    )
    if not fac:
        raise HTTPException(404, f"Faculty '{sec.facultyName}' does not exist")
    sub_ids = [sg.id for sg in fac.subgroups]

    # Payload validation
    if payload.get("id_subgroup") not in sub_ids:
        raise HTTPException(403, "Subgroup does not belong to your faculty")
    if not db.query(Professor).get(payload.get("professor_id")):
        raise HTTPException(404, "Teaching staff member does not exist")

    # → Find current maximum ID
    max_id: int = db.query(func.max(Subject.id)).scalar() or 0
    new_id = max_id + 1

    # Create object with manual ID
    s = Subject(
        id=new_id,
        topic=payload["topic"],
        subgroup_id=payload["id_subgroup"],
        professor_id=payload["professor_id"]
    )

    db.add(s)
    db.commit()
    db.refresh(s)

    return {
        "id":          s.id,
        "topic":       s.topic,
        "id_subgroup": s.subgroup_id,
        "professor_id":  s.professor_id
    }

# ──────────────────────────────────────────────────────────────────────────────
@app.put("/secretariat/api/subjects/{id}")
async def api_update_subject(
    id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 1) Authorization
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    # 2) Find the subject
    s = db.query(Subject).get(id)
    if not s:
        raise HTTPException(404, "Subject does not exist")

    # 3) Check if it belongs to your faculty
    sec = db.query(Secretariat).get(current_user["id"])
    fac = (
        db.query(Faculty)
          .filter(Faculty.longName == sec.facultyName)
          .first()
    )
    sub_ids = [sg.id for sg in fac.subgroups]
    if s.subgroup_id not in sub_ids:
        raise HTTPException(403, "You do not have access to this subject")

    # 4) Apply allowed modifications
    if "topic" in payload:
        s.topic = payload["topic"]
    if "id_subgroup" in payload:
        if payload["id_subgroup"] not in sub_ids:
            raise HTTPException(403, "Subgroup does not belong to your faculty")
        s.subgroup_id = payload["id_subgroup"]
    if "professor_id" in payload:
        if not db.query(Professor).get(payload["professor_id"]):
            raise HTTPException(404, "Teaching staff member does not exist")
        s.professor_id = payload["professor_id"]

    db.commit()
    db.refresh(s)

    return {
        "id":          s.id,
        "topic":       s.topic,
        "id_subgroup": s.subgroup_id,
        "professor_id":  s.professor_id
    }


# ──────────────────────────────────────────────────────────────────────────────
@app.delete("/secretariat/api/subjects/{id}")
async def api_delete_subject(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # 1) Authorization
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    # 2) Find the subject
    s = db.query(Subject).get(id)
    if not s:
        raise HTTPException(404, "Subject does not exist")

    # 3) Check if it belongs to your faculty
    sec = db.query(Secretariat).get(current_user["id"])
    fac = (
        db.query(Faculty)
          .filter(Faculty.longName == sec.facultyName)
          .first()
    )
    sub_ids = [sg.id for sg in fac.subgroups]
    if s.subgroup_id not in sub_ids:
        raise HTTPException(403, "You do not have access to this subject")

    # 4) Delete
    db.delete(s)
    db.commit()

    return JSONResponse(content={"deleted": id}, status_code=200)


@app.get("/secretariat/api/exams", response_model=List[dict])
def list_exams(
    current_user: dict = Depends(get_current_user),
    db: Session       = Depends(get_db),
):
    # 1) Authorize only the secretariat role
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access denied")

    # 2) Find user's Secretariat object
    sec = db.query(Secretariat).filter_by(id=current_user["id"]).first()
    if not sec:
        raise HTTPException(status_code=404, detail="Secretariat not found")
    faculty_short = sec.facultyName

    # 3) Build query: status + joins + faculty filter
    statuses = ["submitted", "accepted"]
    exam_q = (
        db.query(ExamProposal)
          .filter(ExamProposal.status.in_(statuses))
          .join(Subject,   ExamProposal.subject_id == Subject.id)
          .join(Subgroup, Subject.subgroup_id     == Subgroup.id)
          .join(Faculty, Subgroup.facultyId     == Faculty.id)
          .filter(Faculty.longName == sec.facultyName)
    )

    exams = exam_q.all()

    return [
        {
            "id": e.id,
            "subject": {
                "id":    e.subject.id,
                "topic": e.subject.topic,
                "subgroup": {
                    "studyYear":     e.subject.subgroup.studyYear,
                    "groupName":     e.subject.subgroup.groupName,
                    "subgroupIndex": e.subject.subgroup.subgroupIndex
                }
            },
            "room": {
                "id":   e.room.id if e.room else None,
                "name": e.room.name if e.room else None
            },
            "date":        e.date.isoformat(),
            "duration":    e.duration,
            "assistant": {
                "id":        e.assistant.id,
                "firstName": e.assistant.firstName,
                "lastName":  e.assistant.lastName
            } if e.assistant else None,
            "status":       e.status,
            "professor": {
                "id": e.subject.professor.id,
                "firstName": e.subject.professor.firstName,
                "lastName": e.subject.professor.lastName
            } if e.subject.professor else None,
        }
        for e in exams
    ]


# ——————————————————————————————————————————————————————————————————————————————
# MODIFY EXAM — only if status == "submitted"
# ——————————————————————————————————————————————————————————————————————————————
@app.put("/secretariat/api/exams/{exam_id}", response_model=dict)
def update_exam(
    exam_id: int,
    payload: dict,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # 1) Authorize only the secretariat role
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    # 2) Load exam object (and ensure it belongs to the faculty)
    sec = db.query(Secretariat).filter_by(id=current_user["id"]).first()
    exam = (
        db.query(ExamProposal)
          .join(Subject,   ExamProposal.subject_id == Subject.id)
          .join(Subgroup, Subject.subgroup_id     == Subgroup.id)
          .join(Faculty, Subgroup.facultyId     == Faculty.id)
          .filter(ExamProposal.id == exam_id)
          .filter(Faculty.longName == sec.facultyName)
          .first()
    )
    if not exam:
        raise HTTPException(404, "Exam not found or does not belong to your faculty")
    if exam.status not in ("submitted", "accepted"):
        raise HTTPException(
            status_code=400,
            detail="This exam cannot be modified"
        )

    # 3) Clean payload
    allowed = {"date", "room_id", "assistant_id", "professor_id"}
    if "professor_id" in payload:
        professor = (
            db.query(Professor)
            .filter(
                Professor.id == payload["professor_id"],
                Professor.facultyName == sec.facultyName
            )
            .first()
        )

    if not professor:
        raise HTTPException(
            status_code=404,
            detail="Professor not found"
        )

    exam.subject.professor_id = professor.id
    
    for key in list(payload.keys()):
        if key not in allowed:
            payload.pop(key)
    
    # 4) Assign new values
    if "date" in payload:
        exam.date = datetime.fromisoformat(payload["date"])
    if "room_id" in payload:
        exam.room_id = payload["room_id"]
    if "assistant_id" in payload:
        exam.assistant_id = payload["assistant_id"]

    # 5) Save and return updated object
    db.commit()
    db.refresh(exam)
    return {
        "id":           exam.id,
        "subject_id":   exam.subject_id,
        "room_id":      exam.room_id,
        "date":         exam.date.isoformat(),
        "duration":     exam.duration,
        "assistant_id": exam.assistant_id,
        "status":       exam.status,
    }


# ——————————————————————————————————————————————————————————————————————————————
# DELETE EXAM — any, but only from current faculty
# ——————————————————————————————————————————————————————————————————————————————
@app.delete("/secretariat/api/exams/{exam_id}", status_code=204)
def delete_exam(
    exam_id: int,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user["role"] != "secretariat":
        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )
    sec = db.query(Secretariat).filter_by(
        id=current_user["id"]
    ).first()

    if not sec:
        raise HTTPException(
            status_code=404,
            detail="Secretariat not found"
        )
    
    exam = (
        db.query(ExamProposal)
          .join(Subject,   ExamProposal.subject_id == Subject.id)
          .join(Subgroup, Subject.subgroup_id     == Subgroup.id)
          .join(Faculty, Subgroup.facultyId     == Faculty.id)
          .filter(ExamProposal.id == exam_id)
          .filter(Faculty.longName == sec.facultyName)
          .first()
    )
    if not exam:
        raise HTTPException(404, "Exam not found or does not belong to your faculty")
    db.delete(exam)
    db.commit()
    return


@app.get("/secretariat/api/assistants", response_model=list[dict])
def list_assistants(
    current_user: dict = Depends(get_current_user),
    db: Session       = Depends(get_db)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")
    sec = db.query(Secretariat).filter_by(id=current_user["id"]).first()
    fac = db.query(Faculty).filter_by(longName=sec.facultyName).first()
    # All Subgroup + subjects in faculty
    sub_ids = [sg.id for sg in fac.subgroups]
    subject_ids = [s.id for s in db.query(Subject).filter(Subject.subgroup_id.in_(sub_ids)).all()]
    # Professor who are not tenure/lead professors for these subjects
    assistants = (
        db.query(Professor)
          .filter(Professor.id.notin_(
              db.query(Subject.professor_id).filter(Subject.id.in_(subject_ids))
          ))
          .filter(Professor.facultyName == sec.facultyName)
          .all()
    )
    return [
        {"id": a.id, "firstName": a.firstName, "lastName": a.lastName}
        for a in assistants
    ]


@app.get("/secretariat/api/exams-status", response_model=dict)
def list_subjects_with_exam_status(
    current_user: dict = Depends(get_current_user),
    db: Session       = Depends(get_db),
    search:    str    = Query(None, description="filter by subject name or subgroup"),
    page:      int    = Query(1, ge=1, description="page number (starting from 1)"),
    page_size: int    = Query(50, ge=1, le=100, description="items per page"),
):
    # 1) Authorization
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    # 2) Define `sec` before using it
    sec = db.query(Secretariat).filter_by(id=current_user["id"]).first()
    if not sec:
        raise HTTPException(404, "Secretariat not found")
    
    # Subquery for status + sort_key
    status_case = case(
        (ExamProposal.status == None,       "not_submitted"),
        (ExamProposal.status == "accepted", "accepted"),
        (ExamProposal.status == "submitted","submitted"),
        else_="not_submitted"
    ).label("status")
    order_case = case(
        (ExamProposal.status == "accepted", 0),
        (ExamProposal.status == "submitted", 1),
        else_= 2
    )
    
    # Base query on Subject (+ necessary joins)
    q = (
      db.query(
        Subject.id.label("subjectId"),
        Subject.topic.label("topic"),
        Subgroup.groupName.label("groupName"),
        Subgroup.subgroupIndex.label("subgroupIndex"),
        status_case,
        order_case.label("order_key")
      )
      .outerjoin(ExamProposal, ExamProposal.subject_id == Subject.id)
      .join(Subgroup, Subject.subgroup_id  == Subgroup.id)
      .join(Faculty, Subgroup.facultyId  == Faculty.id)
      .filter(Faculty.longName == sec.facultyName)
    )
    
    # Apply search filter if present
    if search:
        term = f"%{search.lower()}%"
        q = q.filter(
          or_(
            Subject.topic.ilike(term),
            Subgroup.groupName.concat(Subgroup.subgroupIndex.cast(String)).ilike(term)
          )
        )
    
    # Calculate total for pagination
    total = q.count()
    
    # Apply sorting and pagination
    items = (
      q.order_by("order_key", Subject.topic)
       .offset((page-1)*page_size)
       .limit(page_size)
       .all()
    )
    
    # Build payload
    data = [
      {
        "subjectId": row.subjectId,
        "topic":     row.topic,
        "subgroup": {
          "groupName":     row.groupName,
          "subgroupIndex": row.subgroupIndex
        },
        "status": row.status
      }
      for row in items
    ]
    
    return {
      "total": total,
      "page": page,
      "page_size": page_size,
      "items": data
    }


def _get_exams_status(db: Session, faculty_long: str):
    # Define CASE expressions for status and sort key
    status_case = case(
        (ExamProposal.status == None,       "not_submitted"),
        (ExamProposal.status == "accepted", "accepted"),
        (ExamProposal.status == "submitted","submitted"),
        else_="not_submitted"
    ).label("status")

    order_case = case(
        (ExamProposal.status == "accepted", 0),
        (ExamProposal.status == "submitted", 1),
        else_=2
    ).label("order_key")

    q = (
        db.query(
            Subject.topic.label("Subject"),
            (Subgroup.groupName + Subgroup.subgroupIndex.cast(String)).label("Subgroup"),
            Room.name.label("Room"),
            ExamProposal.date.label("Date"),
            ExamProposal.duration.label("Duration"),
            (Professor.firstName + " " + Professor.lastName).label("Assistant"),
            status_case,
            order_case
        )
        .outerjoin(ExamProposal, ExamProposal.subject_id == Subject.id)
        .join(Subgroup, Subject.subgroup_id  == Subgroup.id)
        .join(Faculty, Subgroup.facultyId  == Faculty.id)
        .outerjoin(Room,    ExamProposal.room_id     == Room.id)
        .outerjoin(Professor, ExamProposal.assistant_id == Professor.id)
        .filter(Faculty.longName == faculty_long)
        .order_by(order_case, Subject.topic)
    )

    return [row._asdict() for row in q.all()]


@app.get("/secretariat/api/exams/export/excel")
def export_exams_excel(
    current_user: dict = Depends(get_current_user),
    db: Session       = Depends(get_db),
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")

    sec = db.query(Secretariat).filter_by(id=current_user["id"]).first()
    if not sec:
        raise HTTPException(404, "Secretariat not found")

    data = _get_exams_status(db, sec.facultyName)

    df = pd.DataFrame(data)
    if "order_key" in df.columns:
        df = df.drop(columns=["order_key"])
    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exams_status.xlsx"}
    )


@app.get("/secretariat/api/exams/export/pdf")
def export_exams_pdf(
    current_user: dict = Depends(get_current_user),
    db: Session       = Depends(get_db),
):
    if current_user["role"] != "secretariat":
        raise HTTPException(403, "Access denied")
    sec = db.query(Secretariat).filter_by(id=current_user["id"]).first()
    data = _get_exams_status(db, sec.facultyName)

    # Build header list without 'order_key'
    headers = [h for h in data[0].keys() if h != "order_key"] if data else []

    # Prepare wrap style
    styles = getSampleStyleSheet()
    normal_wrap = styles["BodyText"]
    normal_wrap.fontName = "DejaVuSans"
    normal_wrap.fontSize = 8
    normal_wrap.leading = 10

    # Build table_data with Paragraph for text wrapping and without order_key
    table_data = [headers]
    for row in data:
        table_data.append([
            Paragraph(str(row["Subject"]), normal_wrap),
            Paragraph(str(row["Subgroup"]), normal_wrap),
            Paragraph(str(row["Room"] or "-"), normal_wrap),
            Paragraph(str(row["Date"]), normal_wrap),
            Paragraph(str(row["Duration"]), normal_wrap),
            Paragraph(str(row["Assistant"] or "-"), normal_wrap),
            Paragraph(str(row["status"]), normal_wrap),
        ])

    # Adjust colWidths based on column count (7 columns excluding order_key)
    col_widths = [100, 40, 60, 60, 40, 80, 50]

    # Create table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME",      (0, 0), (-1, -1), "DejaVuSans"),
        ("BACKGROUND",    (0, 0), (-1, 0), colors.lightgrey),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.black),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
    ]))

    # Prepare PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=30, rightMargin=30,
        topMargin=30, bottomMargin=30
    )
    doc.build([table])
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=exams_status.pdf"}
    )


@app.get("/professor/assistant-exams/json")
def get_assistant_exams_json(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # 1. Check role
    if current_user["role"] != "professor":
        raise HTTPException(403, "Access denied")

    professor = db.query(Professor).filter_by(emailAddress=current_user["email"]).first()
    if not professor:
        raise HTTPException(404, "professor not found")

    # 2. Gather proposals where assigned as assistant
    proposals = (
        db.query(ExamProposal)
          .join(Subject).join(Subgroup)
          .filter(
            ExamProposal.status == "accepted",
            ExamProposal.assistant_id == professor.id
          )
          .all()
    )

    # 3. Build JSON response
    items_list = []
    for e in proposals:
        sub = e.subject.subgroup
        room = e.room
        items_list.append({
            "id": e.id,
            "subject":       e.subject.topic,            # string
            "studyYear":     sub.studyYear,              # number
            "groupName":     sub.groupName,              # string
            "subgroupIndex": sub.subgroupIndex,          # string
            "date":          e.date.isoformat(),         # ISO string
            "duration":      e.duration,
            "room":          room.name if room else None,
        })

    return JSONResponse({
        "professor": {
          "firstName": professor.firstName,
          "lastName":  professor.lastName
        },
        "exams": items_list
    })


@app.get("/secretariat/exam-limits/json", response_class=JSONResponse)
def get_exam_limits_json(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Access denied")
    limits = db.query(ExamLimits).first()
    if not limits:
        # Return null values
        return {"start_date": None, "end_date": None}
    return {
        "start_date": limits.start_date.isoformat(),
        "end_date":   limits.end_date.isoformat()
    }


@app.get("/group-leader/exam-limits/json", response_class=JSONResponse)
def get_exam_limits_group_leader_json(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "group_leader":
        raise HTTPException(status_code=403, detail="Access denied")
    limits = db.query(ExamLimits).first()
    if not limits:
        # Return null values
        return {"start_date": None, "end_date": None}
    return {
        "start_date": limits.start_date.isoformat(),
        "end_date":   limits.end_date.isoformat()
    }