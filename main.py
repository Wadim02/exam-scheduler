"""
Modulul principal al aplicației ExamScheduler.

Conține toate endpointurile FastAPI, autentificarea Google, procesarea fișierelor Excel și 
rutele pentru dashboard și interfața utilizatorului.
"""
from dotenv import load_dotenv
load_dotenv()
from email_utils import trimite_email_sefi_grupa
from fastapi import FastAPI, Request, Depends, HTTPException, status, UploadFile, File, Form , Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from google.oauth2 import id_token
from google.auth.transport.requests import Request as GoogleRequest
from datetime import datetime, timedelta
from models import Facultati, Cadre, Secretariat, Admin, Sefgrupe, Disciplina, Subgrupe, PropunereExamen, Sali, ExamenLimite
from database import SessionLocal, engine, Base
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import joinedload
import pandas as pd
import jwt
import os
import time
import io
import json
import asyncio
import aiohttp
from io import BytesIO
from typing import Optional
from fastapi import Form
from schemas import PropunereCreate
from datetime import timedelta, timedelta
from sqlalchemy import select, text,create_engine, or_ , and_
from openpyxl import Workbook
import requests as http_requests
from sqlalchemy.dialects.postgresql import insert as pg_insert
from starlette.concurrency import run_in_threadpool
import traceback
from fpdf import FPDF

app = FastAPI()
templates = Jinja2Templates(directory="templates")

# Configurații JWT și Google
SECRET_KEY = os.getenv("SECRET_KEY", "secret-key-puternică")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "client_id_default")

# Creează tabelele în baza de date dacă nu există
Base.metadata.create_all(bind=engine)

async def fetch_json_with_timeout(session, url, timeout_sec=7):
    try:
        async with session.get(url, timeout=timeout_sec) as resp:
            resp.raise_for_status()
            return await resp.json()
    except Exception as e:
        raise e

def get_db():
    """
    Creează o sesiune de DB pentru fiecare request.
    """
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

security = HTTPBearer()

def create_access_token(data: dict, expires_delta: timedelta = None):
    """
    Creează un token JWT pentru autentificare.
    """
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(
    request: Request,
    db: Session = Depends(get_db),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False))
):
    """
    Extrage utilizatorul curent din JWT (din antet sau cookie).
    """
    token = credentials.credentials if credentials else request.cookies.get("access_token")

    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        role = payload.get("role")
        if not email or not role:
            raise HTTPException(status_code=403, detail="Token invalid")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=403, detail="Token expirat")
    except jwt.PyJWTError:
        raise HTTPException(status_code=403, detail="Token invalid sau corupt")
    return {"email": email, "role": role}

@app.post("/token")
async def handle_google_token(request: Request, db: Session = Depends(get_db)):
    """
    Procesează login-ul cu Google și creează token JWT.
    """
    data = await request.json()
    google_token = data.get("token")

    if not google_token:
        raise HTTPException(status_code=400, detail="Token Google lipsă")

    try:
        idinfo = id_token.verify_oauth2_token(
            google_token,
            GoogleRequest(),
            GOOGLE_CLIENT_ID
        )
        email = idinfo["email"]

        # Verifică în toate tabelele cine este utilizatorul
        cadru = db.query(Cadre).filter(Cadre.emailAddress == email).first()
        sec = db.query(Secretariat).filter(Secretariat.emailAddress == email).first()
        admin = db.query(Admin).filter(Admin.emailAddress == email).first()
        sef = db.query(Sefgrupe).filter(Sefgrupe.emailAddress == email).first()

        if cadru:
            role = "cadru"
            full_name = f"{cadru.firstName} {cadru.lastName}"
        elif sec:
            role = "secretariat"
            full_name = f"{sec.firstName} {sec.lastName}"
        elif admin:
            role = "admin"
            full_name = f"{admin.firstName} {admin.lastName}"
        elif sef:
            role = "sef_grupa"
            full_name = f"{sef.firstName} {sef.lastName}"
        else:
            raise HTTPException(status_code=403, detail="Emailul nu este înregistrat în sistem")

        token_data = {
            "sub": email,
            "role": role
        }
        access_token = create_access_token(data=token_data)

        response = RedirectResponse(url="/", status_code=303)
        response.set_cookie(key="access_token", value=access_token, httponly=True)

        return response

    except ValueError as e:
        raise HTTPException(status_code=401, detail=f"Token Google invalid: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare server: {str(e)}")

@app.get("/login")
async def login_page(request: Request):
    """
    Afișează pagina de autentificare.
    """
    return templates.TemplateResponse("login.html", {
        "request": request,
        "google_client_id": GOOGLE_CLIENT_ID
    })

@app.get("/dashboard")
async def dashboard(current_user: dict = Depends(get_current_user)):
    """
    Pagina de dashboard, protejată cu autentificare.
    """
    return {"message": f"Bun venit, {current_user['email']}!"}

@app.get("/users/me")
async def read_users_me(current_user: dict = Depends(get_current_user)):
    """
    Returnează informații despre utilizatorul autentificat.
    """
    return current_user

@app.get("/server-time")
async def get_server_time():
    """
    Returnează timpul curent pe server (UTC și local).
    """
    return {
        "timestamp": int(time.time()),
        "datetime": str(datetime.utcnow()),
        "timezone": str(datetime.now().astimezone().tzinfo)
    }

@app.get("/facultati/")
def list_facultati(db: Session = Depends(get_db)):
    """
    Listează toate facultățile din baza de date.
    """
    return db.query(Facultati).all()

@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: Session = Depends(get_db)):
    """
    Pagina principală a aplicației. Verifică autentificarea și afișează numele utilizatorului.
    """
    token = request.cookies.get("access_token")
    if not token:
        return RedirectResponse(url="/login")

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
    except:
        return RedirectResponse(url="/login")

    user_data = None
    for model in [Cadre, Secretariat, Admin, Sefgrupe]:
        persoana = db.query(model).filter(model.emailAddress == email).first()
        if persoana:
            user_data = {
                "nume_complet": f"{persoana.firstName} {persoana.lastName}",
                "email": email,
                "rol": model.__name__.lower()
            }
            break

    if not user_data:
        return RedirectResponse(url="/login")

    return templates.TemplateResponse("home.html", {
        "request": request,
        "user": user_data,
        "full_name": user_data["nume_complet"]
    })


@app.get("/logout")
async def logout():
    """
    Șterge tokenul și redirecționează la pagina de login.
    """
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("access_token")
    return response

@app.get("/upload-discipline", response_class=HTMLResponse)
async def upload_page(request: Request, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """
    Afișează formularul pentru încărcarea disciplinelor.
    Doar utilizatorii din secretariat au acces.
    """
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    return templates.TemplateResponse("upload_discipline.html", {
        "request": request
    })


@app.get("/propuneri")
async def vezi_discipline_sef_grupa(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    Returnează disciplinele aferente grupei șefului de grupă autentificat.
    """
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces permis doar șefilor de grupă")

    sef = db.query(Sefgrupe).filter(Sefgrupe.emailAddress == current_user["email"]).first()
    if not sef:
        raise HTTPException(status_code=404, detail="Șeful de grupă nu a fost găsit")

    discipline = db.query(Disciplina).filter(Disciplina.grupa == sef.grupa).all()

    return {
        "grupa": sef.grupa,
        "discipline": [
            {
                "id": d.id,
                "nume": d.nume,
                "titular": d.titular,
                "email": d.email
            } for d in discipline
        ]
    }


@app.get("/sefgrupa/discipline")
async def discipline_sef_grupa(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sef = db.query(Sefgrupe).filter(Sefgrupe.emailAddress == current_user["email"]).first()
    if not sef:
        raise HTTPException(status_code=404, detail="Sef de grupă nu găsit")

    subquery = db.query(PropunereExamen.id_disciplina).filter(
        PropunereExamen.status.in_(["trimisa", "acceptata"])
    ).subquery()

    discipline = db.query(Disciplina).filter(
        Disciplina.id_subgrupa == sef.id_subgrupe,
        ~Disciplina.id.in_(select(subquery))
    ).all()

    return [{"id": d.id, "topic": d.topic, "cadru": d.cadru.firstName + " " + d.cadru.lastName} for d in discipline]


@app.get("/sefgrupa/propunere", response_class=HTMLResponse)
async def afiseaza_formular_propunere(request: Request, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sef = db.query(Sefgrupe).filter(Sefgrupe.emailAddress == current_user["email"]).first()

    subquery = db.query(PropunereExamen.id_disciplina).filter(
        PropunereExamen.status.in_(["trimisa", "acceptata"])
    ).subquery()

    discipline = db.query(Disciplina).filter(
        Disciplina.id_subgrupa == sef.id_subgrupe,
        ~Disciplina.id.in_(subquery)
    ).all()

    return templates.TemplateResponse("sefgrupa_propunere.html", {
        "request": request,
        "discipline": discipline
    })


@app.get("/sefgrupa/propunere-form", response_class=HTMLResponse)
async def propunere_form(request: Request, current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sef = db.query(Sefgrupe).filter(Sefgrupe.emailAddress == current_user["email"]).first()

    subquery = db.query(PropunereExamen.id_disciplina).filter(
        PropunereExamen.status.in_(["trimisa", "acceptata"])
    ).subquery()

    discipline = db.query(Disciplina).filter(
        Disciplina.id_subgrupa == sef.id_subgrupe,
        ~Disciplina.id.in_(subquery)
    ).all()

    return templates.TemplateResponse("select_discipline.html", {
        "request": request,
        "discipline": discipline
    })


@app.post("/sefgrupa/propunere")
async def trimite_propunere(
    propunere: PropunereCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sef = db.query(Sefgrupe).filter(Sefgrupe.emailAddress == current_user["email"]).first()

    start = propunere.data.replace(tzinfo=None)
    end = start + timedelta(hours=propunere.durata)

    propuneri_existente = db.query(PropunereExamen).join(Disciplina).filter(
        Disciplina.id_subgrupa == sef.id_subgrupe,
        PropunereExamen.status.in_(["trimisa", "acceptata"])
    ).all()

    for p in propuneri_existente:
        p_start = p.data.replace(tzinfo=None)
        p_end = p_start + timedelta(hours=p.durata)

        if start < p_end and end > p_start:
            raise HTTPException(status_code=409, detail="Intervalul propus se suprapune cu alt examen.")

    prop = PropunereExamen(
        id_disciplina=propunere.disciplina_id,
        id_sefgrupa=sef.id,
        data=propunere.data,
        durata=propunere.durata,
        status="trimisa"
    )
    db.add(prop)
    db.commit()

    return RedirectResponse("/sefgrupa/propunere", status_code=303)

@app.get("/sefgrupa/ocupare-calendar")
async def ocupare_calendar(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sef = db.query(Sefgrupe).filter(Sefgrupe.emailAddress == current_user["email"]).first()
    if not sef:
        raise HTTPException(status_code=404, detail="Șef de grupă nu găsit")

    propuneri = db.query(PropunereExamen).join(Disciplina).filter(
        Disciplina.id_subgrupa == sef.id_subgrupe
    ).all()

    return [
        {
            "data": p.data,
            "durata": p.durata,
            "status": p.status,
            "motiv": getattr(p, "motiv_respingere", None),
            "disciplina": p.disciplina.topic
        }
        for p in propuneri
    ]


@app.get("/sefgrupa/calendar", response_class=HTMLResponse)
async def afiseaza_calendar(
    request: Request,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)  # ← asta lipsea
):
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces interzis")

    limite = db.query(ExamenLimite).first()

    return templates.TemplateResponse("calendar.html", {
        "request": request,
        "start_limit": limite.data_inceput.isoformat() if limite else "",
        "end_limit": limite.data_sfarsit.isoformat() if limite else ""
    })


@app.get("/cadru/propuneri", response_class=HTMLResponse)
async def afiseaza_propuneri_cadru(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "cadru":
        raise HTTPException(status_code=403, detail="Acces permis doar cadrelor didactice")

    cadru = db.query(Cadre).filter(Cadre.emailAddress == current_user["email"]).first()
    if not cadru:
        raise HTTPException(status_code=404, detail="Cadru didactic nu găsit")

    propuneri = db.query(PropunereExamen).join(Disciplina).filter(
        Disciplina.id_cadru == cadru.id,
        PropunereExamen.status == "trimisa"
    ).all()

    toate_sali = db.query(Sali).all()
    colegi = db.query(Cadre).filter(
        Cadre.departmentName == cadru.departmentName,
        Cadre.id != cadru.id
    ).all()

    sali_disponibile = {}
    asistenti_disponibili = {}

    for prop in propuneri:
        start = prop.data.replace(tzinfo=None)
        end = start + timedelta(hours=prop.durata)

        ocupate = db.query(PropunereExamen).filter(
            PropunereExamen.id_sala.isnot(None),
            PropunereExamen.status == "acceptata"
        ).all()

        sali_ocupate_ids = set()
        for p in ocupate:
            p_start = p.data.replace(tzinfo=None)
            p_end = p_start + timedelta(hours=p.durata)
            if start < p_end and end > p_start:
                sali_ocupate_ids.add(p.id_sala)

        sali_disponibile[prop.id] = [s for s in toate_sali if s.id not in sali_ocupate_ids]

        disponibili = []
        for a in colegi:
            ex_propuneri = db.query(PropunereExamen).filter(
                PropunereExamen.id_asistent == a.id,
                PropunereExamen.status == "acceptata"
            ).all()

            conflict = False
            for ep in ex_propuneri:
                ep_start = ep.data.replace(tzinfo=None)
                ep_end = ep_start + timedelta(hours=ep.durata)
                if start < ep_end and end > ep_start:
                    conflict = True
                    break
            if not conflict:
                disponibili.append(a)

        asistenti_disponibili[prop.id] = disponibili

    return templates.TemplateResponse("cadru_propuneri.html", {
        "request": request,
        "propuneri": propuneri,
        "cadru": cadru,
        "sali_disponibile": sali_disponibile,
        "asistenti_disponibili": asistenti_disponibili
    })



@app.post("/cadru/propuneri/accepta")
async def accepta_propunere(
    id_propunere: int = Form(...),
    id_sala: int = Form(...),
    id_asistent: int = Form(...),
    durata: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "cadru":
        raise HTTPException(status_code=403, detail="Acces permis doar cadrelor didactice")

    cadru = db.query(Cadre).filter(Cadre.emailAddress == current_user["email"]).first()
    if not cadru:
        raise HTTPException(status_code=404, detail="Cadru didactic nu găsit")

    prop = db.query(PropunereExamen).filter(PropunereExamen.id == id_propunere).first()
    if not prop:
        raise HTTPException(status_code=404, detail="Propunerea nu există")

    disciplina = db.query(Disciplina).filter(Disciplina.id == prop.id_disciplina).first()
    if not disciplina or disciplina.id_cadru != cadru.id:
        raise HTTPException(status_code=403, detail="Nu ești titularul acestei discipline")

    prop.status = "acceptata"
    prop.id_sala = id_sala
    prop.id_asistent = id_asistent
    prop.durata = durata

    db.commit()

    return RedirectResponse("/cadru/propuneri", status_code=303)


@app.post("/cadru/propuneri/respinge")
async def respinge_propunere(
    id_propunere: int = Form(...),
    motiv: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    cadru = db.query(Cadre).filter(Cadre.emailAddress == current_user["email"]).first()
    prop = db.query(PropunereExamen).filter(PropunereExamen.id == id_propunere).first()

    if not prop:
        raise HTTPException(status_code=404, detail="Propunerea nu a fost găsită")

    disciplina = db.query(Disciplina).filter(Disciplina.id == prop.id_disciplina).first()

    if not disciplina or disciplina.id_cadru != cadru.id:
        raise HTTPException(status_code=403, detail="Nu ai voie să respingi această propunere")

    prop.status = "respinsa"
    prop.motiv_respingere = motiv
    db.commit()

    return RedirectResponse(url="/cadru/propuneri", status_code=303)



@app.post("/cadru/propunere/{id}/refuza")
async def respinge_propunere(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "cadru":
        raise HTTPException(status_code=403, detail="Doar cadrele didactice pot respinge propuneri")

    cadru = db.query(Cadre).filter(Cadre.emailAddress == current_user["email"]).first()
    propunere = db.query(PropunereExamen).filter_by(id=id).first()

    if not propunere:
        raise HTTPException(status_code=404, detail="Propunerea nu există")

    disciplina = db.query(Disciplina).filter_by(id=propunere.id_disciplina).first()
    if not disciplina or disciplina.id_cadru != cadru.id:
        raise HTTPException(status_code=403, detail="Nu ești titularul acestei discipline")

    propunere.status = "respinsa"
    db.commit()

    return RedirectResponse(url="/cadru/propuneri", status_code=303)


@app.get("/cadru/examene-acceptate", response_class=HTMLResponse)
async def examene_acceptate_cadru(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "cadru":
        raise HTTPException(status_code=403, detail="Acces permis doar cadrelor didactice")

    cadru = db.query(Cadre).filter(Cadre.emailAddress == current_user["email"]).first()
    if not cadru:
        raise HTTPException(status_code=404, detail="Cadru didactic nu găsit")

    propuneri = db.query(PropunereExamen).join(Disciplina).filter(
        Disciplina.id_cadru == cadru.id,
        PropunereExamen.status == "acceptata"
    ).all()

    return templates.TemplateResponse("examene_acceptate.html", {
        "request": request,
        "cadru": cadru,
        "examene": propuneri
    })


@app.get("/sefgrupa/discipline-status", response_class=HTMLResponse)
async def status_discipline_sefgrupa(
    request: Request,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sef = db.query(Sefgrupe).filter(Sefgrupe.emailAddress == current_user["email"]).first()
    if not sef:
        raise HTTPException(status_code=404, detail="Șef de grupă nu găsit")

    discipline = db.query(Disciplina).filter(Disciplina.id_subgrupa == sef.id_subgrupe).all()
    propuneri = db.query(PropunereExamen).filter(PropunereExamen.id_sefgrupa == sef.id).all()

    propuneri_map = {p.id_disciplina: p for p in propuneri}

    trimise, acceptate, respinse, netrimise = [], [], [], []

    for d in discipline:
        prop = propuneri_map.get(d.id)
        subgrupa = d.subgrupa

        info = {
            "disciplina": d.topic,
            "an": subgrupa.studyYear,
            "grupa": f"{subgrupa.groupName}{subgrupa.subgroupIndex}",
        }

        if prop:
            info.update({
                "data": prop.data,
                "durata": prop.durata,
                "status": prop.status,
                "motiv": prop.motiv_respingere
            })
            if prop.status == "trimisa":
                trimise.append(info)
            elif prop.status == "acceptata":
                acceptate.append(info)
            elif prop.status == "respinsa":
                respinse.append(info)
        else:
            netrimise.append(info)

    return templates.TemplateResponse("sefgrupa_discipline_status.html", {
        "request": request,
        "trimise": trimise,
        "acceptate": acceptate,
        "respinse": respinse,
        "netrimise": netrimise,
        "sef": sef
    })

@app.get("/sefgrupa/export-excel")
async def export_excel_sefgrupa(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user["role"] != "sef_grupa":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sef = db.query(Sefgrupe).filter_by(emailAddress=current_user["email"]).first()
    if not sef:
        raise HTTPException(status_code=404, detail="Șef de grupă nu găsit")

    discipline = db.query(Disciplina).filter_by(id_subgrupa=sef.id_subgrupe).all()

    propuneri_map = {
        p.id_disciplina: p
        for p in db.query(PropunereExamen).join(Disciplina).filter(Disciplina.id_subgrupa == sef.id_subgrupe).all()
    }

    data = []
    for d in discipline:
        subgrupa = d.subgrupa
        grupa_str = f"{subgrupa.studyYear} {subgrupa.groupName}{subgrupa.subgroupIndex}"
        p = propuneri_map.get(d.id)

        data.append({
            "Disciplina": d.topic,
            "Status": p.status if p else "netrimisa",
            "Data": p.data.strftime("%Y-%m-%d %H:%M") if p else "",
            "Durata (h)": p.durata if p else "",
            "Grupa": grupa_str,
            "Motiv respingere": p.motiv_respingere if p and p.motiv_respingere else "",
            "Sala": p.sala.name if p and p.sala else "",
            "Asistent": f"{p.asistent.firstName} {p.asistent.lastName}" if p and p.asistent else ""
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Examene")

    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=examene_sefgrupa.xlsx"
    })

@app.get("/cadru/asistent-la", response_class=HTMLResponse)
async def examene_asistent(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "cadru":
        raise HTTPException(status_code=403, detail="Acces permis doar cadrelor didactice")

    cadru = db.query(Cadre).filter_by(emailAddress=current_user["email"]).first()
    if not cadru:
        raise HTTPException(status_code=404, detail="Cadru didactic nu găsit")

    propuneri = db.query(PropunereExamen).join(Disciplina).join(Subgrupe).filter(
        PropunereExamen.status == "acceptata",
        PropunereExamen.id_asistent == cadru.id
    ).all()

    return templates.TemplateResponse("asistent_la.html", {
        "request": request,
        "cadru": cadru,
        "examene": propuneri
    })

@app.get("/cadru/export-excel")
async def export_examene_excel(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user["role"] != "cadru":
        raise HTTPException(status_code=403, detail="Acces permis doar cadrelor")

    cadru = db.query(Cadre).filter_by(emailAddress=current_user["email"]).first()
    if not cadru:
        raise HTTPException(status_code=404, detail="Cadru didactic nu găsit")

    wb = Workbook()
    titular_ws = wb.active
    titular_ws.title = "Examene Titular"
    titular_ws.append(["Disciplina", "An", "Grupa", "Data și Ora", "Durata", "Sala", "Asistent"])

    examene_titular = db.query(PropunereExamen).join(Disciplina).filter(
        Disciplina.id_cadru == cadru.id,
        PropunereExamen.status == "acceptata"
    ).all()

    for p in examene_titular:
        sub = p.disciplina.subgrupa
        grupa = f"{sub.studyYear} {sub.groupName}{sub.subgroupIndex}"
        sala = p.sala.name if p.sala else "-"
        asistent = f"{p.asistent.firstName} {p.asistent.lastName}" if p.asistent else "-"
        titular_ws.append([p.disciplina.topic, sub.studyYear, grupa, p.data.strftime("%Y-%m-%d %H:%M"), p.durata, sala, asistent])

    asistent_ws = wb.create_sheet("Examene Asistent")
    asistent_ws.append(["Disciplina", "An", "Grupa", "Data și Ora", "Durata", "Sala", "Tutore"])

    examene_asistent = db.query(PropunereExamen).join(Disciplina).filter(
        PropunereExamen.id_asistent == cadru.id,
        PropunereExamen.status == "acceptata"
    ).all()

    for p in examene_asistent:
        sub = p.disciplina.subgrupa
        grupa = f"{sub.studyYear} {sub.groupName}{sub.subgroupIndex}"
        sala = p.sala.name if p.sala else "-"
        titular = f"{p.disciplina.cadru.firstName} {p.disciplina.cadru.lastName}"
        asistent_ws.append([p.disciplina.topic, sub.studyYear, grupa, p.data.strftime("%Y-%m-%d %H:%M"), p.durata, sala, titular])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=examene_cadru.xlsx"}
    )

@app.get("/admin/facultati", response_class=HTMLResponse)
async def lista_facultati(request: Request, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    facultati = db.query(Facultati).all()
    return templates.TemplateResponse("admin_facultati.html", {
        "request": request,
        "facultati": facultati
    })

@app.post("/admin/facultati/update")
async def update_facultati(request: Request, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    form = await request.form()
    facultati = db.query(Facultati).all()

    for f in facultati:
        nume_nou = form.get(f"nume_{f.id}")
        short_nou = form.get(f"short_{f.id}")
        if nume_nou and short_nou:
            f.longName = nume_nou.strip()
            f.shortName = short_nou.strip()

    db.commit()
    return RedirectResponse("/admin/facultati", status_code=303)

@app.post("/admin/facultati/add")
async def adauga_facultate(request: Request, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    form = await request.form()
    name = form.get("longName", "").strip()
    short = form.get("shortName", "").strip()

    if not name or not short:
        raise HTTPException(status_code=400, detail="Date lipsă")

    db.execute(text("SELECT setval('facultati_id_seq', (SELECT MAX(id) FROM facultati))"))
    noua = Facultati(longName=name, shortName=short)
    db.add(noua)
    db.commit()

    return RedirectResponse("/admin/facultati", status_code=303)

@app.post("/admin/facultati/delete")
async def sterge_facultate(request: Request, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    form = await request.form()
    id_fac = form.get("id")

    facultate = db.query(Facultati).filter(Facultati.id == id_fac).first()
    if not facultate:
        raise HTTPException(status_code=404, detail="Facultatea nu a fost găsită")

    subgrupe_existente = db.query(Subgrupe).filter(Subgrupe.facultyId == id_fac).count()
    if subgrupe_existente > 0:
        raise HTTPException(status_code=400, detail="Nu poți șterge această facultate deoarece există subgrupe asociate")

    db.delete(facultate)
    db.commit()
    return RedirectResponse("/admin/facultati", status_code=303)

@app.post("/admin/sincronizeaza-seventa-facultati")
async def sincronizeaza_seventa_facultati(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces permis doar administratorilor")

    try:
        db.execute(text("SELECT setval('facultati_id_seq', (SELECT COALESCE(MAX(id), 1) FROM facultati))"))
        db.commit()
        return {"message": "Secvența pentru facultăți a fost sincronizată cu succes!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare la sincronizare: {str(e)}")

@app.get("/admin/cadre", response_class=HTMLResponse)
async def admin_cadre_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    cadre_usm = db.query(Cadre).filter(Cadre.emailAddress.ilike('%@usm.ro')).all()

    return templates.TemplateResponse("admin_cadre.html", {
        "request": request,
        "cadre": cadre_usm
    })


@app.post("/admin/cadre/update")
async def update_cadru(
    id: int = Form(...),
    firstName: str = Form(...),
    lastName: str = Form(...),
    emailAddress: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    cadru = db.query(Cadre).filter(Cadre.id == id).first()
    if not cadru:
        raise HTTPException(status_code=404, detail="Cadru nu a fost găsit")

    cadru.firstName = firstName.strip()
    cadru.lastName = lastName.strip()
    cadru.emailAddress = emailAddress.strip()
    cadru.departmentName = departmentName.strip()
    db.commit()

    return RedirectResponse("/admin/cadre", status_code=303)

@app.post("/admin/cadre/delete/{id}")
async def delete_cadru(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    cadru = db.query(Cadre).filter(Cadre.id == id).first()
    if not cadru:
        raise HTTPException(status_code=404, detail="Cadru nu a fost găsit")

    db.delete(cadru)
    db.commit()

    return RedirectResponse("/admin/cadre", status_code=303)

@app.post("/admin/cadre/create")
async def adauga_cadru(
    firstName: str = Form(...),
    lastName: str = Form(...),
    emailAddress: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Doar adminul poate adăuga cadre")

    if not emailAddress.endswith("@usm.ro"):
        raise HTTPException(status_code=400, detail="Emailul trebuie să fie @usm.ro")

    if db.query(Cadre).filter(Cadre.emailAddress == emailAddress).first():
        raise HTTPException(status_code=400, detail="Emailul este deja folosit")

    db.execute(text("SELECT setval('cadre_id_seq', (SELECT COALESCE(MAX(id), 1) FROM cadre))"))

    nou_cadru = Cadre(
        firstName=firstName.strip(),
        lastName=lastName.strip(),
        emailAddress=emailAddress.strip(),
        departmentName=departmentName.strip()
    )
    db.add(nou_cadru)
    db.commit()
    return RedirectResponse("/admin/cadre", status_code=303)


@app.get("/admin/adauga-cadru", response_class=HTMLResponse)
async def adauga_cadru_form(request: Request, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    facultati = db.query(Facultati).all()
    domenii = db.query(Cadre.departmentName).distinct().all()
    domenii = [d[0] for d in domenii if d[0]]  # extragem doar valorile

    return templates.TemplateResponse("adauga_cadru.html", {
        "request": request,
        "facultati": facultati,
        "domenii": domenii
    })

@app.post("/admin/adauga-cadru")
async def adauga_cadru(
    request: Request,
    lastName: str = Form(...),
    firstName: str = Form(...),
    emailAddress: str = Form(...),
    phoneNumber: Optional[str] = Form(None),
    facultyName: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    # Validare simplă pentru email
    if not emailAddress.endswith("@usm.ro"):
        raise HTTPException(status_code=400, detail="Emailul trebuie să fie de forma @usm.ro")

    # Verificăm dacă deja există un cadru cu acest email
    if db.query(Cadre).filter_by(emailAddress=emailAddress).first():
        raise HTTPException(status_code=400, detail="Emailul există deja în sistem")

    nou = Cadre(
        lastName=lastName.strip(),
        firstName=firstName.strip(),
        emailAddress=emailAddress.strip().lower(),
        phoneNumber=phoneNumber.strip() if phoneNumber else None,
        facultyName=facultyName.strip(),
        departmentName=departmentName.strip()
    )

    db.add(nou)
    db.commit()
    return RedirectResponse("/admin/cadre", status_code=303)


@app.get("/api/departamente")
async def get_departamente(facultate: str, db: Session = Depends(get_db)):
    """
    Returnează lista unică de departamente pentru o facultate dată.
    """
    departamente = db.query(Cadre.departmentName).filter(
        Cadre.facultyName == facultate
    ).distinct().all()
    return [d[0] for d in departamente if d[0]]

@app.get("/admin/secretariat", response_class=HTMLResponse)
async def admin_secretariat_page(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    # Verificare rol
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    # Obține lista tuturor secretarilor
    secretari = db.query(Secretariat).order_by(Secretariat.lastName, Secretariat.firstName).all()

    # Returnează template-ul
    return templates.TemplateResponse("admin_secretariat.html", {
        "request": request,
        "secretari": secretari
    })
@app.post("/admin/secretariat/add")
async def adauga_secretar(
    firstName: str = Form(...),
    lastName: str = Form(...),
    emailAddress: str = Form(...),
    facultyName: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    if db.query(Secretariat).filter_by(emailAddress=emailAddress).first():
        raise HTTPException(status_code=400, detail="Emailul există deja")

    nou = Secretariat(
        firstName=firstName.strip(),
        lastName=lastName.strip(),
        emailAddress=emailAddress.strip().lower(),
        facultyName=facultyName.strip(),
        departmentName=departmentName.strip()
    )

    db.add(nou)
    db.commit()
    return RedirectResponse("/admin/secretariat", status_code=303)


@app.post("/admin/secretariat/update")
async def editeaza_secretar(
    id: int = Form(...),
    firstName: str = Form(...),
    lastName: str = Form(...),
    emailAddress: str = Form(...),
    facultyName: str = Form(...),
    departmentName: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    s = db.query(Secretariat).filter_by(id=id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Secretar inexistent")

    s.firstName = firstName.strip()
    s.lastName = lastName.strip()
    s.emailAddress = emailAddress.strip().lower()
    s.facultyName = facultyName.strip()
    s.departmentName = departmentName.strip()

    db.commit()
    return RedirectResponse("/admin/secretariat", status_code=303)


@app.post("/admin/secretariat/delete")
async def sterge_secretar(
    id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acces interzis")

    s = db.query(Secretariat).filter_by(id=id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Secretar inexistent")

    db.delete(s)
    db.commit()
    return RedirectResponse("/admin/secretariat", status_code=303)


# Endpoint GET pentru descarcare sali + generare fisier sali.xlsx
@app.get("/secretariat/download-sali-status")
async def descarca_si_converteste_sali(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    url = "https://orar.usv.ro/orar/vizualizare/data/sali.php?json"
    folder = "vizualizare_date"
    output_folder = "csv_output"
    os.makedirs(folder, exist_ok=True)
    os.makedirs(output_folder, exist_ok=True)

    filename = "sali.xlsx"
    excel_path = os.path.join(folder, filename)
    csv_filename = "sali.csv"
    csv_path = os.path.join(output_folder, csv_filename)

    try:
        # Descărcare
        response = http_requests.get(url)
        response.raise_for_status()
        data = response.json()

        df = pd.DataFrame(data)
        df = df[df['name'].notna() & (df['name'].str.strip() != "")]
        df = df.sort_values(by='name')

        df.to_excel(excel_path, index=False)

        # Conversie CSV
        df.to_csv(csv_path, index=False)

        return {
            "status": "success",
            "mesaje": [
                "📥 Fișierul sali.xlsx a fost descărcat cu succes.",
                "🔄 Fișierul a fost convertit în sali.csv."
            ],
            "download_url": "/secretariat/descarca-sali-excel"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare la descărcare sau conversie: {str(e)}")
@app.get("/secretariat/descarca-sali-excel")
async def descarca_fisier_excel():
    file_path = os.path.join("vizualizare_date", "sali.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Fișierul nu a fost găsit")
    
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="sali.xlsx")

@app.get("/secretariat/descopera-sali")
async def descarca_sali_si_salveaza_excel(
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    url = "https://orar.usv.ro/orar/vizualizare/data/sali.php?json"
    folder = "vizualizare_date"
    filename = "sali.xlsx"
    output_file = os.path.join(folder, filename)

    try:
        response = http_requests.get(url)
        response.raise_for_status()
        data = response.json()

        df = pd.DataFrame(data)
        df = df[df['name'].notna() & (df['name'].str.strip() != "")]
        df = df.sort_values(by='name')

        os.makedirs(folder, exist_ok=True)
        df.to_excel(output_file, index=False)

        return FileResponse(output_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            filename="sali.xlsx")

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare la descărcare: {str(e)}")


# Endpoint POST pentru incarcarea fisierului modificat si salvarea in DB
@app.post("/secretariat/incarca-sali")
async def incarca_sali_din_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    try:
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))

        # Verificam coloanele esentiale
        if 'name' not in df.columns:
            raise HTTPException(status_code=400, detail="Fișierul Excel trebuie să conțină coloana 'name'")

        db.query(Sali).delete()  # Ștergem tot

        for _, row in df.iterrows():
            sala = Sali(name=row["name"].strip(), building=row.get("building", "").strip())
            db.add(sala)

        db.commit()
        return RedirectResponse("/secretariat", status_code=303)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare la încărcarea fișierului: {str(e)}")

@app.get("/secretariat", response_class=HTMLResponse)
async def secretariat_dashboard(
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    return templates.TemplateResponse("secretariat_sali.html", {"request": request})

@app.post("/secretariat/import-sali-csv")
async def import_sali_csv(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    csv_path = "csv_output/sali.csv"
    if not os.path.exists(csv_path):
        raise HTTPException(status_code=404, detail="Fișierul sali.csv nu există")

    try:
        df = pd.read_csv(csv_path)
        count = 0

        for _, row in df.iterrows():
            stmt = pg_insert(Sali.__table__).values(
                id=int(row["id"]),
                name=str(row.get("name", "")).strip(),
                shortName=str(row.get("shortName", "")).strip(),
                buildingName=str(row.get("buildingName", "")).strip()
            ).on_conflict_do_update(
                index_elements=["id"],  # Dacă există deja acel id...
                set_={
                    "name": str(row.get("name", "")).strip(),
                    "shortName": str(row.get("shortName", "")).strip(),
                    "buildingName": str(row.get("buildingName", "")).strip()
                }
            )
            db.execute(stmt)
            count += 1

        db.commit()
        return {"status": "success", "importate": count}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare la importul fișierului sali.csv: {str(e)}")
@app.get("/secretariat/sali", response_class=HTMLResponse)
async def vizualizeaza_sali(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sali = db.query(Sali).order_by(Sali.id).all()
    return templates.TemplateResponse("secretariat_vizualizare_sali.html", {
        "request": request,
        "sali": sali
    })
    
@app.post("/secretariat/sali/update")
async def actualizeaza_sali(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    form = await request.form()
    sali = db.query(Sali).all()

    for sala in sali:
        sala.name = form.get(f"name_{sala.id}", "").strip()
        sala.shortName = form.get(f"shortName_{sala.id}", "").strip()
        sala.buildingName = form.get(f"buildingName_{sala.id}", "").strip()

    db.commit()
    return RedirectResponse("/secretariat/sali", status_code=303)
@app.post("/secretariat/sali/delete/{id}")
async def sterge_sala(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sala = db.query(Sali).filter_by(id=id).first()
    if not sala:
        raise HTTPException(status_code=404, detail="Sala nu a fost găsită")

    db.delete(sala)
    db.commit()
    return RedirectResponse("/secretariat/sali", status_code=303)
@app.post("/secretariat/sali/add")
async def adauga_sala(
    name: str = Form(...),
    shortName: str = Form(""),
    buildingName: str = Form(""),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    noua = Sali(
        name=name.strip(),
        shortName=shortName.strip(),
        buildingName=buildingName.strip()
    )
    db.execute(text("SELECT setval('sali_id_seq', (SELECT MAX(id) FROM sali))"))
    db.add(noua)
    db.commit()

    return RedirectResponse("/secretariat/sali", status_code=303)
@app.get("/secretariat/sefgrupe", response_class=HTMLResponse)
async def afiseaza_sefgrupe(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sefi = db.query(Sefgrupe).order_by(Sefgrupe.lastName, Sefgrupe.firstName).all()
    subgrupe = db.query(Subgrupe).all()
    facultati = db.query(Facultati).all()

    return templates.TemplateResponse("secretariat_sefgrupe.html", {
        "request": request,
        "sefi": sefi,
        "subgrupe": subgrupe,
        "facultati": facultati
    })

@app.post("/secretariat/sefgrupe/add")
async def adauga_sefgrupa(
    request: Request,
    lastName: str = Form(...),
    firstName: str = Form(...),
    emailAddress: str = Form(...),
    phoneNumber: str = Form(None),
    id_facultate: int = Form(...),
    id_subgrupe: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    # Verificăm dacă există deja un șef cu acest email
    if db.query(Sefgrupe).filter_by(emailAddress=emailAddress.strip().lower()).first():
        return templates.TemplateResponse("secretariat_sefgrupe.html", {
            "request": request,
            "sefi": db.query(Sefgrupe).all(),
            "facultati": db.query(Facultati).all(),
            "subgrupe": db.query(Subgrupe).all(),
            "error_message": "⚠️ Emailul există deja în sistem. Încearcă altul."
        })

    sef_nou = Sefgrupe(
        lastName=lastName.strip(),
        firstName=firstName.strip(),
        emailAddress=emailAddress.strip().lower(),
        phoneNumber=phoneNumber.strip() if phoneNumber else None,
        id_facultate=id_facultate,
        id_subgrupe=id_subgrupe
    )

    db.add(sef_nou)
    db.commit()

    return templates.TemplateResponse("secretariat_sefgrupe.html", {
        "request": request,
        "sefi": db.query(Sefgrupe).all(),
        "facultati": db.query(Facultati).all(),
        "subgrupe": db.query(Subgrupe).all(),
        "success_message": "✅ Șeful de grupă a fost adăugat cu succes!"
    })
    return RedirectResponse("/secretariat/sefgrupe", status_code=303)

@app.post("/secretariat/sefgrupe/update")
async def actualizeaza_sefi_grupa(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    form = await request.form()
    try:
        sef_id = int(form.get("id"))
    except:
        raise HTTPException(status_code=400, detail="ID invalid")

    sef = db.query(Sefgrupe).filter_by(id=sef_id).first()
    if not sef:
        raise HTTPException(status_code=404, detail="Șef de grupă nu a fost găsit")

    # Extragem câmpurile din formular
    sef.firstName = form.get(f"firstName_{sef_id}", sef.firstName).strip()
    sef.lastName = form.get(f"lastName_{sef_id}", sef.lastName).strip()
    sef.emailAddress = form.get(f"emailAddress_{sef_id}", sef.emailAddress).strip()
    sef.phoneNumber = form.get(f"phoneNumber_{sef_id}", sef.phoneNumber).strip()

    try:
        sef.id_facultate = int(form.get(f"id_facultate_{sef_id}"))
        sef.id_subgrupe = int(form.get(f"id_subgrupe_{sef_id}"))
    except ValueError:
        raise HTTPException(status_code=400, detail="ID invalid pentru facultate sau subgrupă")

    db.commit()
    return RedirectResponse("/secretariat/sefgrupe?msg=salvat", status_code=303)

@app.post("/secretariat/sefgrupe/delete/{id}")
async def sterge_sefgrupa(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    sef = db.query(Sefgrupe).filter_by(id=id).first()
    if not sef:
        raise HTTPException(status_code=404, detail="Șef de grupă inexistent")

    # Ștergem propunerile legate de acest șef
    db.query(PropunereExamen).filter_by(id_sefgrupa=id).delete()

    # Acum putem șterge șeful
    db.delete(sef)
    db.commit()

    return RedirectResponse("/secretariat/sefgrupe", status_code=303)

@app.post("/secretariat/sefgrupe/import")
async def importa_sefi_grupa(
    excel_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    try:
        contents = await excel_file.read()
        df = pd.read_excel(BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail="Fișier invalid sau corupt")

    try:
        output_csv = os.path.join("csv_output", "sefgrupe_import.csv")
        df.to_csv(output_csv, index=False)

        for _, row in df.iterrows():
            if db.query(Sefgrupe).filter_by(emailAddress=row["emailAddress"]).first():
                continue  # Skip dacă emailul există
            sef = Sefgrupe(
                lastName=row["lastName"],
                firstName=row["firstName"],
                emailAddress=row["emailAddress"],
                phoneNumber=row["phoneNumber"],
                id_facultate=row["id_facultate"],
                id_subgrupe=row["id_subgrupe"]
            )
            db.add(sef)

        db.commit()
        return {"message": f"{len(df)} șefi de grupă importați cu succes."}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Eroare la procesare: {str(e)}")

@app.post("/secretariat/discipline/genereaza")
async def genereaza_discipline_endpoint(
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    excel_path = "vizualizare_date/subgrupe.xlsx"
    df_ids = pd.read_excel(excel_path)
    id_subgrupe = df_ids["id"].dropna().astype(int).tolist()
    data = []

    async with aiohttp.ClientSession() as session:
        for id_subgrupa in id_subgrupe:
            try:
                url = f"https://orar.usv.ro/orar/vizualizare/data/orarSPG.php?ID={id_subgrupa}&mod=grupa&json"
                json_data = await asyncio.wait_for(fetch_json_with_timeout(session, url), timeout=10)

                if isinstance(json_data, list) and len(json_data) > 0 and isinstance(json_data[0], list):
                    for cls in json_data[0]:
                        if isinstance(cls, dict) and cls.get("typeShortName") == "curs":
                            data.append({
                                "id_cadru": int(cls.get("teacherID", 0)),
                                "id_subgrupa": id_subgrupa,
                                "topic": cls.get("topicLongName", "")
                            })

            except:
                continue  # Ignorăm ID-urile care dau eroare

    df = pd.DataFrame(data)
    df_unique = df.drop_duplicates(subset=["id_cadru", "id_subgrupa", "topic"]).reset_index(drop=True)

    engine = create_engine("postgresql://postgres:ad12min34@localhost:5432/exam_scheduler")

    with engine.connect() as conn:
        result = conn.execute(text("SELECT id FROM cadre"))
        cadre_existente = {row[0] for row in result.fetchall()}

    df_filtrat = df_unique[df_unique["id_cadru"].isin(cadre_existente)].reset_index(drop=True)
    df_filtrat["id"] = df_filtrat.index + 1

    os.makedirs("vizualizare_date", exist_ok=True)
    os.makedirs("csv_output", exist_ok=True)

    df_filtrat.to_excel("vizualizare_date/discipline.xlsx", index=False)
    df_filtrat.to_csv("csv_output/discipline.csv", index=False)

    return {"message": f"{len(df_filtrat)} discipline generate", "csv": "csv_output/discipline.csv"}
@app.post("/secretariat/import-discipline-csv")
async def import_discipline_csv(
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    csv_path = "csv_output/discipline.csv"
    if not os.path.exists(csv_path):
        raise HTTPException(status_code=404, detail="Fișierul discipline.csv nu a fost găsit")

    try:
        df = pd.read_csv(csv_path)

        engine = create_engine("postgresql://postgres:ad12min34@localhost:5432/exam_scheduler")

        with engine.begin() as conn:
            conn.execute(text("DELETE FROM propuneri_examene;"))  # resetăm propunerile
            conn.execute(text("DELETE FROM discipline;"))         # curățăm tabela discipline

        df.to_sql("discipline", engine, index=False, if_exists="append")

        return {"message": f"{len(df)} discipline importate cu succes"}

    except Exception as e:
        return HTTPException(status_code=500, detail=f"Eroare la import: {str(e)}")
@app.get("/secretariat/discipline-admin", response_class=HTMLResponse)
async def admin_discipline_page(
    request: Request,
    page: int = 1,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    per_page = 50
    total = db.query(Disciplina).count()
    discipline = db.query(Disciplina).offset((page - 1) * per_page).limit(per_page).all()
    cadre = db.query(Cadre).all()
    subgrupe = db.query(Subgrupe).all()

    return templates.TemplateResponse("discipline_admin.html", {
        "request": request,
        "discipline": discipline,
        "cadre": cadre,
        "subgrupe": subgrupe,
        "pagina_curenta": page,
        "pagini_totale": (total // per_page) + (1 if total % per_page else 0)
    })

@app.post("/secretariat/discipline/add")
async def adauga_disciplina(
    topic: str = Form(...),
    id_cadru: int = Form(...),
    id_subgrupa: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    noua = Disciplina(topic=topic, id_cadru=id_cadru, id_subgrupa=id_subgrupa)
    db.add(noua)
    db.commit()
    return RedirectResponse("/secretariat/discipline-admin", status_code=303)

@app.post("/secretariat/discipline/update/{id}")
async def update_disciplina(
    id: int,
    topic: str = Form(...),
    id_cadru: int = Form(...),
    id_subgrupa: int = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    d = db.query(Disciplina).filter_by(id=id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Disciplina nu a fost găsită")

    d.topic = topic
    d.id_cadru = id_cadru
    d.id_subgrupa = id_subgrupa
    db.commit()
    return RedirectResponse("/secretariat/discipline-admin", status_code=303)

@app.get("/secretariat/discipline/delete/{id}")
async def sterge_disciplina(
    id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    disciplina = db.query(Disciplina).filter_by(id=id).first()
    if disciplina:
        db.delete(disciplina)
        db.commit()
    return RedirectResponse("/secretariat/discipline-admin", status_code=303)
@app.get("/secretariat/limite-examene", response_class=HTMLResponse)
async def formular_limite_examene(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    limite = db.query(ExamenLimite).first()
    return templates.TemplateResponse("limite_examene.html", {
        "request": request,
        "limite": limite
    })
@app.post("/secretariat/limite-examene")
async def seteaza_limite_examene(
    data_inceput: str = Form(...),
    data_sfarsit: str = Form(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    di = datetime.fromisoformat(data_inceput)
    ds = datetime.fromisoformat(data_sfarsit)

    if di >= ds:
        raise HTTPException(status_code=400, detail="Data de început trebuie să fie înainte de data de sfârșit.")

    limite = db.query(ExamenLimite).first()
    if limite:
        limite.data_inceput = di
        limite.data_sfarsit = ds
    else:
        limite = ExamenLimite(data_inceput=di, data_sfarsit=ds)
        db.add(limite)

    db.commit()
    return RedirectResponse("/secretariat/limite-examene", status_code=303)

@app.get("/secretariat/status-examene")
def status_examene(
    request: Request,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=500)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    # Obține toate disciplinele
    discipline = db.execute(text("""
        SELECT id, id_cadru, id_subgrupa, topic
        FROM discipline
        ORDER BY id
    """)).fetchall()

    # Obține ID-urile disciplinelor cu examen acceptat
    examene_acceptate = db.execute(text("""
        SELECT id_disciplina
        FROM propuneri_examene
        WHERE status = 'acceptata'
    """)).fetchall()
    discipline_programate = {row[0] for row in examene_acceptate}

    rezultat = []
    count_programate = 0

    for d in discipline:
        este_programata = d[0] in discipline_programate
        if este_programata:
            count_programate += 1

        rezultat.append({
            "id": d[0],
            "cadru": d[1],
            "subgrupa": d[2],
            "disciplina": d[3],
            "stare": "programat" if este_programata else "neprogramat"
        })

    total = len(rezultat)
    start = (page - 1) * size
    end = start + size
    paginat = rezultat[start:end]
    total_pagini = (total + size - 1) // size

    return {
        "total": total,
        "programate": count_programate,
        "pagina": page,
        "pe_pagina": size,
        "total_pagini": total_pagini,
        "discipline": paginat
    }

@app.get("/secretariat/status")
def pagina_status_examene(request: Request, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")
    return templates.TemplateResponse("secretariat_status_examene.html", {"request": request})
@app.get("/secretariat/examene", response_class=HTMLResponse)
def lista_examene_programate(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    examene = db.query(PropunereExamen).filter(
        PropunereExamen.status == "acceptata"
    ).options(
        joinedload(PropunereExamen.disciplina),
        joinedload(PropunereExamen.sala),
        joinedload(PropunereExamen.asistent)
    ).all()

    return templates.TemplateResponse("secretariat_examene_lista.html", {
        "request": request,
        "examene": examene
    })

@app.get("/api/examene-programate")
def get_examene_programate(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    examene = db.execute(text("""
        SELECT id, id_disciplina, id_sala, data, durata, id_asistent
        FROM propuneri_examene
        WHERE status = 'acceptata'
        ORDER BY data
    """)).fetchall()

    return [dict(row._mapping) for row in examene]

@app.delete("/api/examene-programate/{examen_id}")
def delete_examen(examen_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    db.execute(text("DELETE FROM propuneri_examene WHERE id = :id AND status = 'acceptata'"), {"id": examen_id})
    db.commit()
    return {"message": "Examenul a fost șters"}

@app.put("/api/examene-programate/{examen_id}")
def update_examen(examen_id: int, payload: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    db.execute(text("""
        UPDATE propuneri_examene
        SET id_sala = :id_sala,
            data = :data,
            durata = :durata,
            id_asistent = :id_asistent
        WHERE id = :id AND status = 'acceptata'
    """), {
        "id": examen_id,
        "id_sala": payload.get("id_sala"),
        "data": payload.get("data"),
        "durata": payload.get("durata"),
        "id_asistent": payload.get("id_asistent")
    })
    db.commit()
    return {"message": "Examenul a fost actualizat"}

@app.post("/api/examene-programate")
def create_examen(payload: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    db.execute(text("""
        INSERT INTO propuneri_examene (id_disciplina, id_sala, data, durata, id_asistent, status)
        VALUES (:id_disciplina, :id_sala, :data, :durata, :id_asistent, 'acceptata')
    """), payload)
    db.commit()
    return {"message": "Examen adăugat"}

@app.get("/secretariat/examen/{id}", response_class=HTMLResponse)
def editeaza_examen(
    id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    examen = db.query(PropunereExamen).filter_by(id=id, status="acceptata").first()
    if not examen:
        raise HTTPException(status_code=404, detail="Examenul nu a fost găsit")

    # Obține sălile disponibile (fără alt examen la aceeași oră)
    toate_sali = db.query(Sali).all()
    sali_ocupate = db.query(PropunereExamen.id_sala).filter(
        PropunereExamen.data == examen.data,
        PropunereExamen.id != id,
        PropunereExamen.status == "acceptata"
    ).all()
    sali_blocate = {s.id_sala for s in sali_ocupate}
    sali_disponibile = [s for s in toate_sali if s.id not in sali_blocate]

    # Cadre din aceeași facultate (asistenți)
    cadre = db.query(Cadre).filter(Cadre.departmentName == examen.disciplina.cadru.departmentName).all()

    return templates.TemplateResponse("secretariat_examen_edit.html", {
        "request": request,
        "examen": examen,
        "sali": sali_disponibile,
        "asistenti": cadre
    })
@app.post("/secretariat/examen/{id}", response_class=HTMLResponse)
def salveaza_examen(
    id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
    data: str = Form(...),
    durata: int = Form(...),
    id_sala: int = Form(...),
    id_asistent: Optional[int] = Form(None)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    examen = db.query(PropunereExamen).filter_by(id=id, status="acceptata").first()
    if not examen:
        return templates.TemplateResponse("secretariat_examen_edit.html", {
            "request": request,
            "examen": examen,
            "sali": [],
            "asistenti": [],
            "mesaj_eroare": "Examenul nu a fost găsit"
        }, status_code=404)

    start_nou = datetime.fromisoformat(data)
    end_nou = start_nou + timedelta(minutes=durata)

    # Verifică dacă sala e disponibilă în acel interval
    examene_in_sala = db.query(PropunereExamen).filter(
        PropunereExamen.id != id,
        PropunereExamen.status == "acceptata",
        PropunereExamen.id_sala == id_sala
    ).all()
    for e in examene_in_sala:
        e_start = e.data
        e_end = e_start + timedelta(minutes=e.durata)
        if start_nou < e_end and end_nou > e_start:
            # Regenerăm contextul paginii
            toate_sali = db.query(Sali).all()
            sali_ocupate = db.query(PropunereExamen.id_sala).filter(
                PropunereExamen.data == examen.data,
                PropunereExamen.id != id,
                PropunereExamen.status == "acceptata"
            ).all()
            sali_blocate = {s.id_sala for s in sali_ocupate}
            sali_disponibile = [s for s in toate_sali if s.id not in sali_blocate]
            cadre = db.query(Cadre).filter(Cadre.departmentName == examen.disciplina.cadru.departmentName).all()

            return templates.TemplateResponse("secretariat_examen_edit.html", {
                "request": request,
                "examen": examen,
                "sali": sali_disponibile,
                "asistenti": cadre,
                "mesaj_eroare": "Sala selectată este deja ocupată în acel interval."
            }, status_code=200)
    # Verifică disponibilitatea asistentului
    if id_asistent:
        propuneri_conflict = db.query(PropunereExamen).join(Disciplina).filter(
            PropunereExamen.id != id,
            PropunereExamen.status == "acceptata",
            or_(
                PropunereExamen.id_asistent == id_asistent,
                Disciplina.id_cadru == id_asistent
            )
        ).all()

        for p in propuneri_conflict:
            p_start = p.data
            p_end = p_start + timedelta(minutes=p.durata)
            if start_nou < p_end and end_nou > p_start:
                # Refacem datele necesare paginii
                toate_sali = db.query(Sali).all()
                sali_ocupate = db.query(PropunereExamen.id_sala).filter(
                    PropunereExamen.data == examen.data,
                    PropunereExamen.id != id,
                    PropunereExamen.status == "acceptata"
                ).all()
                sali_blocate = {s.id_sala for s in sali_ocupate}
                sali_disponibile = [s for s in toate_sali if s.id not in sali_blocate]
                cadre = db.query(Cadre).filter(Cadre.departmentName == examen.disciplina.cadru.departmentName).all()

                return templates.TemplateResponse("secretariat_examen_edit.html", {
                    "request": request,
                    "examen": examen,
                    "sali": sali_disponibile,
                    "asistenti": cadre,
                    "mesaj_eroare": "Asistentul selectat este deja ocupat în acel interval."
                }, status_code=200)

    # Salvare
    examen.data = start_nou
    examen.durata = durata
    examen.id_sala = id_sala
    examen.id_asistent = id_asistent if id_asistent else None

    db.commit()
    return RedirectResponse(url="/secretariat/examene", status_code=303)

@app.get("/secretariat/sali-disponibile")
def sali_disponibile(
    data: str,
    durata: int = 90,
    examen_id: int = 0,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    try:
        start = datetime.fromisoformat(data)
    except ValueError:
        raise HTTPException(status_code=400, detail="Format dată invalid")

    end = start + timedelta(minutes=durata)

    examene_conflict = db.query(PropunereExamen).filter(
        PropunereExamen.status == "acceptata",
        PropunereExamen.id != examen_id
    ).all()

    sali_ocupate = set()
    for ex in examene_conflict:
        ex_start = ex.data
        ex_end = ex_start + timedelta(minutes=ex.durata)
        if start < ex_end and end > ex_start:
            sali_ocupate.add(ex.id_sala)

    sali_libere = db.query(Sali).filter(~Sali.id.in_(sali_ocupate)).all()

    return [{"id": s.id, "nume": s.nume} for s in sali_libere]

@app.get("/secretariat/planificare/excel")
def export_planificare_excel(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    examene = db.query(PropunereExamen).filter(
        PropunereExamen.status == "acceptata"
    ).options(
        joinedload(PropunereExamen.disciplina).joinedload(Disciplina.cadru),
        joinedload(PropunereExamen.sala),
        joinedload(PropunereExamen.asistent)
    ).all()

    # ✅ Sortare după semigrupă
    examene.sort(
        key=lambda e: f"{e.disciplina.subgrupa.groupName}{e.disciplina.subgrupa.subgroupIndex}"
    )

    rows = []
    for e in examene:
        semigrupa = f"{e.disciplina.subgrupa.groupName}{e.disciplina.subgrupa.subgroupIndex}"
        an = e.disciplina.subgrupa.studyYear or "—"

        rows.append({
            "Disciplina": e.disciplina.topic,
            "Titular": f"{e.disciplina.cadru.firstName} {e.disciplina.cadru.lastName}",
            "Data examen": e.data.strftime("%Y-%m-%d %H:%M"),
            "Durata (minute)": e.durata,
            "Sala": e.sala.name if e.sala else "—",
            "Asistent": f"{e.asistent.firstName} {e.asistent.lastName}" if e.asistent else "—",
            "Semigrupa": semigrupa,
            "An": an
        })

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Planificare")

    output.seek(0)
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=planificare_examene.xlsx"}
    )

@app.get("/secretariat/planificare/pdf")
def export_planificare_pdf(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "secretariat":
        raise HTTPException(status_code=403, detail="Acces interzis")

    examene = db.query(PropunereExamen).filter(
        PropunereExamen.status == "acceptata"
    ).options(
        joinedload(PropunereExamen.disciplina).joinedload(Disciplina.cadru),
        joinedload(PropunereExamen.sala),
        joinedload(PropunereExamen.asistent)
    ).all()

    # ✅ Sortare după semigrupă (ex: 3111A, 3111B, etc.)
    examene.sort(
        key=lambda e: f"{e.disciplina.subgrupa.groupName}{e.disciplina.subgrupa.subgroupIndex}"
    )

    pdf = FPDF(orientation='L', unit='mm', format='A4')  # landscape
    pdf.add_page()
    pdf.add_font("DejaVu", "", "static/fonts/DejaVuSans.ttf", uni=True)
    pdf.add_font("DejaVu", "B", "static/fonts/DejaVuSans-Bold.ttf", uni=True)
    pdf.set_font("DejaVu", size=10)

    pdf.set_fill_color(200, 220, 255)
    pdf.cell(0, 10, "Planificare examene", ln=True, align="C")

    # Headerele tabelului
    headers = ["Disciplina", "Titular", "Data", "Durata", "Sala", "Asistent", "Semigrupa", "An"]
    col_widths = [50, 40, 30, 18, 25, 40, 30, 15]
    pdf.set_font("DejaVu", size=9)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, 1, 0, "C", fill=True)
    pdf.ln()

    last_semigrupa = None
    # Rândurile tabelului
    for e in examene:
        semigrupa = f"{e.disciplina.subgrupa.groupName}{e.disciplina.subgrupa.subgroupIndex}"
        an = str(e.disciplina.subgrupa.studyYear) if e.disciplina.subgrupa.studyYear else "—"

        # 🟦 Adaugă titlu pentru noua semigrupă
        if semigrupa != last_semigrupa:
            pdf.set_font("DejaVu", style="B", size=10)
            pdf.set_fill_color(230, 230, 250)  # mov deschis
            pdf.cell(sum(col_widths), 8, f"Semigrupa: {semigrupa} | An: {an}", 1, ln=1, fill=True)
            pdf.set_font("DejaVu", size=9)
            last_semigrupa = semigrupa

        values = [
            e.disciplina.topic,
            f"{e.disciplina.cadru.firstName} {e.disciplina.cadru.lastName}",
            e.data.strftime("%Y-%m-%d %H:%M"),
            f"{e.durata} min",
            e.sala.name if e.sala else "—",
            f"{e.asistent.firstName} {e.asistent.lastName}" if e.asistent else "—",
            semigrupa,
            an
        ]

        for i, val in enumerate(values):
            pdf.cell(col_widths[i], 10, str(val), 1, 0, "L")
        pdf.ln()

    pdf_bytes = pdf.output(dest="S").encode("latin1")

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=planificare_examene.pdf"}
    )